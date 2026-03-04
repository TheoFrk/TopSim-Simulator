"""
TOPSIM Adaptive Intelligence System V6.1 "Eagle Eye"
Selbstlernender Simulator mit automatischer Kalibrierung.
V5.4: Auto-Fit (linear/power/quad/log), Optimierer, Handbuch-konform.
"""

import json
import os
import math
import glob as globmod
from copy import deepcopy

import numpy as np
import optuna
optuna.logging.set_verbosity(optuna.logging.WARNING)
from scipy.optimize import least_squares, minimize, differential_evolution


# ---------------------------------------------------------------------------
# Auto-Fit: Erkennung Zusammenhangstyp
# ---------------------------------------------------------------------------

MODEL_TYPES = ("power", "linear", "quadratic", "logarithmic")


def _fit_all_models(x_vals, y_vals):
    """Testet 4 Modelltypen und gibt den besten zurueck.
    Returns: (typ, params_dict, r2, predict_fn)"""
    x = np.array(x_vals, dtype=float)
    y = np.array(y_vals, dtype=float)
    n = len(x)
    if n < 2:
        return "power", {"a": 1.0, "b": 1.0}, 0.0, lambda v: v

    ss_tot = np.sum((y - np.mean(y)) ** 2)
    if ss_tot < 1e-12:
        return "constant", {"c": np.mean(y)}, 1.0, lambda v: np.mean(y)

    results = []

    # 1) Linear: y = a*x + b
    try:
        A = np.vstack([x, np.ones(n)]).T
        coef, *_ = np.linalg.lstsq(A, y, rcond=None)
        pred = coef[0] * x + coef[1]
        r2 = 1 - np.sum((y - pred) ** 2) / ss_tot
        results.append(("linear", {"a": coef[0], "b": coef[1]}, r2,
                         lambda v, a=coef[0], b=coef[1]: a * v + b))
    except (np.linalg.LinAlgError, ValueError, FloatingPointError, OverflowError):
        pass

    # 2) Quadratic: y = a*x^2 + b*x + c
    try:
        A = np.vstack([x**2, x, np.ones(n)]).T
        coef, *_ = np.linalg.lstsq(A, y, rcond=None)
        pred = coef[0] * x**2 + coef[1] * x + coef[2]
        r2 = 1 - np.sum((y - pred) ** 2) / ss_tot
        results.append(("quadratic", {"a": coef[0], "b": coef[1], "c": coef[2]}, r2,
                         lambda v, a=coef[0], b=coef[1], c=coef[2]: a * v**2 + b * v + c))
    except (np.linalg.LinAlgError, ValueError, FloatingPointError, OverflowError):
        pass

    # 3) Power: y = a * x^b  (log-linear fit)
    try:
        mask = (x > 0) & (y > 0)
        if mask.sum() >= 2:
            lx, ly = np.log(x[mask]), np.log(y[mask])
            A = np.vstack([lx, np.ones(mask.sum())]).T
            coef, *_ = np.linalg.lstsq(A, ly, rcond=None)
            b_val, a_log = coef
            a_val = math.exp(a_log)
            pred = a_val * x[mask] ** b_val
            r2 = 1 - np.sum((y[mask] - pred) ** 2) / np.sum((y[mask] - np.mean(y[mask])) ** 2)
            results.append(("power", {"a": a_val, "b": b_val}, r2,
                             lambda v, a=a_val, b=b_val: a * v ** b))
    except (np.linalg.LinAlgError, ValueError, FloatingPointError, OverflowError, ZeroDivisionError):
        pass

    # 4) Logarithmic: y = a * ln(x) + b
    try:
        mask = x > 0
        if mask.sum() >= 2:
            lx = np.log(x[mask])
            A = np.vstack([lx, np.ones(mask.sum())]).T
            coef, *_ = np.linalg.lstsq(A, y[mask], rcond=None)
            pred = coef[0] * lx + coef[1]
            r2 = 1 - np.sum((y[mask] - pred) ** 2) / np.sum((y[mask] - np.mean(y[mask])) ** 2)
            results.append(("logarithmic", {"a": coef[0], "b": coef[1]}, r2,
                             lambda v, a=coef[0], b=coef[1]: a * math.log(max(v, 1e-9)) + b))
    except (np.linalg.LinAlgError, ValueError, FloatingPointError, OverflowError, ZeroDivisionError):
        pass

    if not results:
        return "power", {"a": 1.0, "b": 1.0}, 0.0, lambda v: v

    best = max(results, key=lambda r: r[2])
    return best


def _model_label(typ, params):
    if typ == "linear":
        return f"y = {params['a']:.4f}*x + {params['b']:.2f}"
    elif typ == "quadratic":
        return f"y = {params['a']:.6f}*x² + {params['b']:.4f}*x + {params['c']:.2f}"
    elif typ == "power":
        return f"y = {params['a']:.4f} * x^{params['b']:.4f}"
    elif typ == "logarithmic":
        return f"y = {params['a']:.4f}*ln(x) + {params['b']:.2f}"
    return str(params)

# ---------------------------------------------------------------------------
# Excel-Parser  (xlrd fuer .xls, openpyxl fuer .xlsx)
# ---------------------------------------------------------------------------

def _open_workbook(path):
    if path.endswith(".xls"):
        import xlrd
        return _XlrdWrapper(xlrd.open_workbook(path))
    else:
        import openpyxl
        return _OpenpyxlWrapper(openpyxl.load_workbook(path, data_only=True))


class _XlrdWrapper:
    def __init__(self, wb):
        self._wb = wb
    def sheet_names(self):
        return self._wb.sheet_names()
    def sheet(self, name):
        ws = self._wb.sheet_by_name(name)
        return [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]


class _OpenpyxlWrapper:
    def __init__(self, wb):
        self._wb = wb
    def sheet_names(self):
        return self._wb.sheetnames
    def sheet(self, name):
        ws = self._wb[name]
        return [[v if v is not None else "" for v in row] for row in ws.iter_rows(values_only=True)]


def _safe_float(val, default=0.0):
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _prompt_float(prompt, default=None, min_val=None, max_val=None):
    while True:
        raw = input(prompt).strip()
        if raw == "" and default is not None:
            val = float(default)
        else:
            try:
                val = float(raw)
            except ValueError:
                print("  Ungueltige Zahl. Bitte erneut eingeben.")
                continue
        if min_val is not None and val < min_val:
            print(f"  Wert zu klein (min {min_val}).")
            continue
        if max_val is not None and val > max_val:
            print(f"  Wert zu gross (max {max_val}).")
            continue
        return val


def _prompt_int(prompt, default=None, min_val=None, max_val=None):
    while True:
        raw = input(prompt).strip()
        if raw == "" and default is not None:
            val = int(default)
        else:
            try:
                val = int(raw)
            except ValueError:
                print("  Ungueltige Ganzzahl. Bitte erneut eingeben.")
                continue
        if min_val is not None and val < min_val:
            print(f"  Wert zu klein (min {min_val}).")
            continue
        if max_val is not None and val > max_val:
            print(f"  Wert zu gross (max {max_val}).")
            continue
        return val


def _prompt_yes_no(prompt, default=False):
    default_label = "j" if default else "n"
    while True:
        raw = input(f"{prompt} [{default_label}]: ").strip().lower()
        if raw == "":
            return default
        if raw in {"j", "ja", "y", "yes", "1", "true"}:
            return True
        if raw in {"n", "nein", "no", "0", "false"}:
            return False
        print("  Bitte 'j' oder 'n' eingeben.")


# ---------------------------------------------------------------------------
# Report-Parser
# ---------------------------------------------------------------------------

class ReportParser:
    def __init__(self, path, unternehmen_nr=2):
        self.path = path
        self.wb = _open_workbook(path)
        self.u_nr = unternehmen_nr

    def _find_u_col(self, header_row, target=None):
        if target is None:
            target = f"U{self.u_nr}"
        for i, v in enumerate(header_row):
            if str(v).strip() == target:
                return i
        return None

    def _get_period_col(self, rows):
        header = rows[3]
        last_col = len(header) - 1
        while last_col > 0 and header[last_col] == "":
            last_col -= 1
        return last_col

    def _row_val(self, rows, label_substr, col, exact=False):
        for row in rows:
            if not row:
                continue
            cell = str(row[0]).strip()
            if exact and cell == label_substr:
                return _safe_float(row[col])
            if not exact and label_substr.lower() in cell.lower():
                return _safe_float(row[col])
        return None

    @staticmethod
    def _first_numeric(row, start=1):
        for i in range(start, len(row)):
            if isinstance(row[i], (int, float)) and row[i] != 0:
                return row[i]
        return 0.0

    def parse(self):
        data = {"_source": self.path}
        self._parse_covering(data)
        self._parse_executive(data)
        self._parse_markt(data)
        self._parse_fertigung(data)
        self._parse_fe(data)
        self._parse_lager(data)
        self._parse_personal(data)
        self._parse_deckungsbeitrag(data)
        self._parse_guv(data)
        self._parse_bilanz(data)
        self._parse_wertorientiert(data)
        self._parse_entscheidungen(data)
        return data

    def _parse_covering(self, data):
        rows = self.wb.sheet("Covering Page")
        for row in rows:
            txt = str(row[0]).strip()
            if txt.startswith("Periode:"):
                data["periode"] = int(txt.split(":")[1].strip())
                return
        data["periode"] = 0

    def _parse_executive(self, data):
        rows = self.wb.sheet("1) Executive Summary")
        col = self._get_period_col(rows)
        mapping = {
            "aktienkurs": "Aktienkurs",
            "umsatz_gesamt": "Umsatz Gesamt",
            "periodenueberschuss": "Periodenüberschuss",
            "absatz_gesamt": "Absatz Gesamt",
            "tats_absatz_m1": "Tatsächlicher Absatz",
            "marktanteil": "Marktanteil",
            "fertigungsmenge_geplant": "Geplante Fertigungsmenge",
            "fertigungsmenge_tats": "Tatsächliche Fertigungsmenge",
            "lager_fertig": "Lagerendbestand",
            "auslastung_ma": "Auslastung Mitarbeiter",
            "auslastung_anlagen": "Auslastung Anlagen",
            "herstellkosten_stueck": "Herstellkosten",
            "selbstkosten_stueck": "Selbstkosten",
            "db5_stueck": "Deckungsbeitrag V",
            "betriebsergebnis": "Betriebsergebnis",
            "eigenkapital": "Eigenkapital",
            "eigenkapitalrendite": "Eigenkapitalrendite",
            "umsatzrendite": "Umsatzrendite",
            "kassenendbestand": "Kassenendbestand",
            "ueberziehungskredit": "Überziehungskredit",
            "finanzergebnis": "Finanzergebnis",
            "fremdkapitalquote": "Fremdkapitalquote",
        }
        for key, label in mapping.items():
            val = self._row_val(rows, label, col)
            if val is not None:
                data[key] = val

        for row in rows:
            label = str(row[0]).strip() if row else ""
            if label == "Rating":
                val = row[col] if col < len(row) else ""
                if isinstance(val, str) and val.strip():
                    data["rating"] = val.strip()

    def _parse_markt(self, data):
        try:
            rows = self.wb.sheet("2) Marktforschungsbericht")
        except (KeyError, Exception) as e:
            print(f"  WARNUNG: Tabellenblatt 'Markt' nicht gefunden oder fehlerhaft: {e}")
            return
        u_col = self._find_u_col(rows[3])
        if u_col is None:
            u_col = 3

        mapping = {
            "preis": "Preis",
            "tech_index": "Technologie",
            "werbung": "Werbung",
            "vertrieb_ma": "Vertriebsmitarbeiter",
            "kz_index": "Kundenzufriedenheit",
            "bekanntheit": "Bekanntheit",
            "pot_absatz": "Potentieller Absatz",
            "tats_absatz_markt": "Tatsächlicher Absatz",
            "nicht_gedeckt": "Nicht gedeckte Nachfrage",
            "umsatz_markt": "Umsatz Markt",
            "marktanteil_markt": "Marktanteil",
        }
        for key, label in mapping.items():
            val = self._row_val(rows, label, u_col)
            if val is not None:
                data[key] = val

        all_competitors = {}
        for u_num in range(1, 7):
            c = self._find_u_col(rows[3], f"U{u_num}")
            if c is None:
                continue
            comp = {}
            for key, label in mapping.items():
                val = self._row_val(rows, label, c)
                if val is not None:
                    comp[key] = val
            all_competitors[f"U{u_num}"] = comp
        data["alle_unternehmen"] = all_competitors

        avg_col = self._find_u_col(rows[3], "ø-Wert / Summe")
        if avg_col is None:
            for i, v in enumerate(rows[3]):
                if "Summe" in str(v) or "ø" in str(v):
                    avg_col = i
                    break

        branche_preise = [c.get("preis", 0) for c in all_competitors.values() if c.get("preis", 0) > 0]
        data["branche_avg_preis"] = sum(branche_preise) / len(branche_preise) if branche_preise else 3000
        branche_kz = [c.get("kz_index", 60) for c in all_competitors.values() if c.get("kz_index", 0) > 0]
        data["branche_avg_kz"] = sum(branche_kz) / len(branche_kz) if branche_kz else 60.0
        data["branche_nicht_gedeckt_summe"] = sum(
            c.get("nicht_gedeckt", 0) for c in all_competitors.values()
        )
        data["branche_pot_absatz_summe"] = sum(
            c.get("pot_absatz", 0) for c in all_competitors.values()
        )
        data["branche_tats_absatz_summe"] = sum(
            c.get("tats_absatz_markt", 0) for c in all_competitors.values()
        )

        for row in rows:
            if "Großabnehmer" in str(row[0]):
                data["grossabnehmer"] = _safe_float(row[u_col])
                break

    def _parse_fertigung(self, data):
        try:
            rows = self.wb.sheet("3) Fertigungsbericht")
        except (KeyError, Exception) as e:
            print(f"  WARNUNG: Tabellenblatt 'Fertigung' nicht gefunden oder fehlerhaft: {e}")
            return
        for row in rows:
            label = str(row[0]).strip()
            if "Kapazität der Anlagen" in label and "ohne" in label:
                data["anlagen_kapazitaet"] = self._first_numeric(row, 2)
            if "Verfügbare Fertigungskapazität" in label:
                data["anlagen_kap_mit_ueberstunden"] = self._first_numeric(row, 2)
            if label == "Summe" and len(row) > 5 and _safe_float(row[4]) > 0:
                data["abschreibungen_summe"] = _safe_float(row[4])
                data["anlagen_restbuchwert"] = _safe_float(row[5])
                data["anlagen_fixkosten"] = _safe_float(row[6])
            if label == "Summe" and len(row) > 2:
                kap = _safe_float(row[1])
                umwelt = _safe_float(row[2])
                if 10000 < kap < 200000:
                    data["anlagen_kapazitaet_summe"] = kap
                if 50 < umwelt < 120:
                    data["umweltindex_anlagen"] = umwelt
            if "Produktivitätsindex I" in label:
                data["produktivitaet_index1"] = self._first_numeric(row, 1)
            if "Produktivitätsindex II" in label:
                data["produktivitaet_index2"] = self._first_numeric(row, 1)
            if "Mitarbeitermotivation" in label:
                data["motivation"] = self._first_numeric(row, 1)
            if "Fehlzeiten" in label and "Anzahl" not in label:
                pass
            if label == "- Fehlzeiten":
                data["fehlzeiten"] = self._first_numeric(row, 1)
            if "Einsetzbares Personal" in label and "ohne" in label:
                data["einsetzbares_personal"] = self._first_numeric(row, 1)
            if "Kumulierte Fertigung" in label:
                data["kumul_fertigung"] = self._first_numeric(row, 1)

    def _parse_fe(self, data):
        try:
            rows = self.wb.sheet("4) Forschung & Entwicklung")
        except (KeyError, Exception) as e:
            print(f"  WARNUNG: Tabellenblatt '4) Forschung & Entwicklung' nicht gefunden oder fehlerhaft: {e}")
            return
        g1_found = False
        g2_found = False
        for row in rows:
            row_str = " ".join(str(v) for v in row)
            if not g1_found and ("Gen. 1" in row_str or "Gen 1" in row_str):
                nums = [v for v in row if isinstance(v, (int, float))]
                if len(nums) >= 2:
                    data["fe_invest_gen1"] = float(nums[0])
                    data["tech_index_result"] = float(nums[1])
                    g1_found = True
            if not g2_found and ("Gen. 2" in row_str or "Gen 2" in row_str):
                nums = [v for v in row if isinstance(v, (int, float))]
                if len(nums) >= 2:
                    data["fe_invest_gen2"] = float(nums[0])
                    data["tech_index_gen2"] = float(nums[1])
                    g2_found = True

    def _parse_lager(self, data):
        try:
            rows = self.wb.sheet("5) Lager")
        except (KeyError, Exception) as e:
            print(f"  WARNUNG: Tabellenblatt 'Lager' nicht gefunden oder fehlerhaft: {e}")
            return
        for row in rows:
            label = str(row[0]).strip()
            if label == "Lagerendbestand Fertigerzeugnisse":
                data["lager_fertig"] = _safe_float(row[2])
                break

    def _parse_personal(self, data):
        try:
            rows = self.wb.sheet("6) Personal")
        except (KeyError, Exception) as e:
            print(f"  WARNUNG: Tabellenblatt 'Personal' nicht gefunden oder fehlerhaft: {e}")
            return
        for row in rows:
            label = str(row[0]).strip()
            if label == "Personalanfangsbestand":
                data["personal_fertigung"] = _safe_float(row[4])
                data["personal_fe"] = _safe_float(row[5])
                data["personal_vertrieb"] = _safe_float(row[6])
                data["personal_summe"] = _safe_float(row[7])
            if label == "Personalendbestand":
                data["personal_fert_end"] = _safe_float(row[4])
                data["personal_fe_end"] = _safe_float(row[5])
                data["personal_vertrieb_end"] = _safe_float(row[6])

    def _parse_deckungsbeitrag(self, data):
        try:
            rows = self.wb.sheet("10) Deckungsbeitragsrechnung")
        except (KeyError, Exception) as e:
            print(f"  WARNUNG: Tabellenblatt 'Deckungsbeitrag' nicht gefunden oder fehlerhaft: {e}")
            return
        for row in rows:
            label = str(row[0]).strip()
            if label == "Umsatzerlöse" and isinstance(row[1], (int, float)):
                data["umsatz_m1"] = _safe_float(row[1])
                data["umsatz_gross"] = _safe_float(row[2])
            if "Variable Materialkosten" in label and isinstance(row[-1], (int, float)):
                data["var_materialkosten"] = _safe_float(row[-1])
            if "Variable Fertigungskosten" in label and isinstance(row[-1], (int, float)):
                data["var_fertigungskosten"] = _safe_float(row[-1])
            if "Transportkosten" in label and isinstance(row[-1], (int, float)):
                data["transportkosten"] = _safe_float(row[-1])
            if label == "= Deckungsbeitrag I" and isinstance(row[-1], (int, float)):
                data["db1"] = _safe_float(row[-1])
            if label == "= Deckungsbeitrag V" and isinstance(row[-1], (int, float)):
                data["db5"] = _safe_float(row[-1])

    def _parse_guv(self, data):
        try:
            rows = self.wb.sheet("11) Gewinn- und Verlustrechnung")
        except (KeyError, Exception) as e:
            print(f"  WARNUNG: Tabellenblatt 'GuV' nicht gefunden oder fehlerhaft: {e}")
            return
        for row in rows:
            label = str(row[0]).strip()
            if label == "- Materialaufwand":
                data["materialaufwand"] = _safe_float(row[1])
            if label == "- Personalaufwand":
                data["personalaufwand"] = _safe_float(row[1])
            if label == "- Löhne/Gehälter":
                data["loehne"] = _safe_float(row[1])
            if label == "- Abschreibungen":
                data["abschreibungen"] = _safe_float(row[1])
            if label == "- Sonstiger Aufwand":
                data["sonstiger_aufwand"] = _safe_float(row[1])
            if label == "= Betriebsergebnis":
                data["ebit"] = _safe_float(row[1])
            if label == "= Gewinn vor Steuern":
                data["gewinn_vor_steuern"] = _safe_float(row[1])
            if label.startswith("- Steuern"):
                data["steuern"] = _safe_float(row[1])
            if "Erhöhung/Verminderung" in label:
                data["bestandsveraenderung_guv"] = _safe_float(row[1])

    def _parse_bilanz(self, data):
        try:
            rows = self.wb.sheet("14) Bilanz")
        except (KeyError, Exception) as e:
            print(f"  WARNUNG: Tabellenblatt 'Bilanz' nicht gefunden oder fehlerhaft: {e}")
            return
        for row in rows:
            label = str(row[0]).strip()
            if label == "Anlagevermögen":
                data["anlagevermoegen"] = _safe_float(row[1])
            if label == "Bilanzsumme":
                data["bilanzsumme"] = _safe_float(row[1])
            if "Fertige Erzeugnisse" in label:
                data["lager_wert"] = _safe_float(row[1])
            if "Forderungen" in label:
                data["forderungen"] = _safe_float(row[1])
            passiv_label = str(row[3]).strip() if len(row) > 3 else ""
            if passiv_label == "Eigenkapital":
                data["eigenkapital"] = _safe_float(row[4])
            if "Überziehungskredit" in passiv_label:
                data["ueberziehungskredit"] = _safe_float(row[4])
            if "Pensionsrückstellungen" in passiv_label:
                data["pensionsrueckstellungen"] = _safe_float(row[4])

    def _parse_wertorientiert(self, data):
        try:
            rows = self.wb.sheet("16) Wertorientierte Kennzahlen")
        except (KeyError, Exception) as e:
            print(f"  WARNUNG: Tabellenblatt 'Wertorientiert' nicht gefunden oder fehlerhaft: {e}")
            return
        col = self._get_period_col(rows)
        for key, label in [
            ("eigenkapitalquote_wk", "Eigenkapitalquote"),
            ("wacc", "WACC"), ("nopat", "NOPAT"),
            ("cashflow", "Traditioneller Cashflow"),
            ("nce", "Net Capital Employed"),
            ("eva", "Economic Value Added (EVA"),
            ("mva", "Market Value Added"),
            ("ebit_wk", "EBIT"), ("ebitda", "EBITDA"),
            ("free_cashflow", "Free Cashflow"),
            ("fremdkapital_wk", "Fremdkapital"),
        ]:
            val = self._row_val(rows, label, col)
            if val is not None:
                data[key] = val

    def _parse_entscheidungen(self, data):
        try:
            rows = self.wb.sheet("17) Entscheidungsprotokoll")
        except (KeyError, Exception) as e:
            print(f"  WARNUNG: Tabellenblatt '17) Entscheidungsprotokoll' nicht gefunden oder fehlerhaft: {e}")
            return

        def _last_numeric(row):
            for i in range(len(row) - 1, 0, -1):
                if isinstance(row[i], (int, float)):
                    return row[i]
            return 0.0

        col = self._get_period_col(rows)
        decisions = {}
        label_map = {
            "Preis Markt 1": "preis_m1", "Werbung Markt 1": "werbung_m1",
            "Preis Markt 2": "preis_m2", "Werbung Markt 2": "werbung_m2",
            "Vertrieb Personalendbestand": "vertrieb_ma",
            "Technologie Personalendbestand": "tech_personal",
            "Großabnehmer": "grossabnehmer", "Ausschreibung": "ausschreibung_preis",
            "Fertigungsmenge": "fertigungsmenge",
        }
        for row in rows:
            label = str(row[0]).strip()
            row_str = " ".join(str(v) for v in row)
            for search, key in label_map.items():
                if search in label:
                    val = _safe_float(row[col]) if col < len(row) and isinstance(row[col], (int, float)) else 0.0
                    if val == 0.0:
                        val = _last_numeric(row)
                    decisions[key] = val
            if "Investition" in label and "Typ A" in label:
                decisions["neue_anlagen_a"] = _safe_float(row[col]) if col < len(row) else 0.0
            elif "Einstellungen/Entlassungen" in label:
                decisions["personal_aenderung_fert"] = _last_numeric(row)
            elif "Gen. 2" in row_str and "Technologie" in row_str and "Personal" in row_str:
                nums = [v for v in row if isinstance(v, (int, float))]
                data["personal_fe_gen2"] = float(nums[0]) if nums else 0.0
            elif "Technologie Personal" in row_str or ("Technologie" in row_str and "Personalendbestand" in row_str):
                nums = [v for v in row if isinstance(v, (int, float))]
                data["personal_fe"] = float(nums[0]) if nums else 35.0
                decisions["personal_fe"] = data["personal_fe"]
            elif "Vertrieb" in row_str and "Anz" in row_str:
                nums = [v for v in row if isinstance(v, (int, float))]
                data["vertrieb_ma"] = int(nums[0]) if len(nums) > 0 else data.get("vertrieb_ma", 100)
                data["vertrieb_m2"] = int(nums[1]) if len(nums) > 1 else 0
            elif "Einsatzstoffe" in row_str or "Teile" in row_str:
                nums = [v for v in row if isinstance(v, (int, float))]
                data["einkauf_menge"] = float(nums[0]) if nums else 45000.0
            elif "Marktforschungsbericht" in row_str:
                # Checkbox: prüfe ob eine 1, True oder "x" in der Zeile steht
                data["marktforschung_aktiv"] = any(v in (1, True, "x", "X", "ja", "Ja") for v in row)
            elif "Ökologie" in row_str and "Gen. 1" in row_str:
                nums = [v for v in row if isinstance(v, (int, float))]
                data["oeko_budget_gen1"] = float(nums[0]) if nums else 0.0
            elif "Ökologie" in row_str and "Gen. 2" in row_str:
                nums = [v for v in row if isinstance(v, (int, float))]
                data["oeko_budget_gen2"] = float(nums[0]) if nums else 0.0
            elif "Corporate Identity" in row_str:
                nums = [v for v in row if isinstance(v, (int, float))]
                data["ci_budget"] = float(nums[0]) if nums else 0.0
            elif "Großabnehmer" in row_str:
                nums = [v for v in row if isinstance(v, (int, float))]
                data["grossabnehmer"] = int(nums[0]) if nums else 0
            elif "Desinvestitionen" in row_str or "Anlage Nr." in row_str:
                data["desinvest"] = []
        data["decisions"] = decisions


# ---------------------------------------------------------------------------
# History Database
# ---------------------------------------------------------------------------

class HistoryDB:
    def __init__(self, path="history_v6.json"):
        self.path = path
        self.data = self._load()

    def _load(self):
        if os.path.exists(self.path):
            with open(self.path) as f:
                return json.load(f)
        return {"perioden": {}}

    def save(self):
        with open(self.path, "w") as f:
            json.dump(self.data, f, indent=2, ensure_ascii=False)

    def add_period(self, period_nr, report_data):
        self.data["perioden"][str(period_nr)] = report_data
        self.save()

    def get_period(self, nr):
        return self.data["perioden"].get(str(nr))

    def all_periods(self):
        return sorted(self.data["perioden"].keys(), key=int)

    def period_count(self):
        return len(self.data["perioden"])


# ---------------------------------------------------------------------------
# Calibration Engine
# ---------------------------------------------------------------------------

RATING_ZINS_MAP = {
    "AAA": -4.0, "AA": -3.0, "A": -2.0, "BBB": -1.0, "BB": 0.0,
    "B": 1.0, "CCC": 2.0, "CC": 3.0, "C": 4.0, "D": 5.0,
}

UMWELT_STRAF_STAFFEL = [
    (85.0, 5.0), (90.0, 2.5), (92.5, 1.5), (95.0, 1.0), (98.0, 0.5), (100.0, 0.0),
]


def umwelt_strafe(index):
    if index >= 100.0:
        return 0.0
    for grenze, strafe in UMWELT_STRAF_STAFFEL:
        if index <= grenze:
            return strafe
    return 0.0


def rating_zinssatz(rating_str, basiszins=8.0):
    delta = RATING_ZINS_MAP.get(rating_str, 0.0)
    return max(0.5, basiszins + delta)


VERWALTUNG_KURVE = [
    (50, 150), (75, 170), (100, 190), (125, 200), (150, 215),
    (175, 228), (200, 240), (250, 260),
]

EINKAUF_KURVE = [
    (50, 10), (75, 14), (100, 16), (125, 18), (150, 20),
    (175, 21), (200, 22), (250, 24),
]


def _interpolate_kurve(kurve, umsatz_meur):
    if umsatz_meur <= kurve[0][0]:
        return kurve[0][1]
    for i in range(len(kurve) - 1):
        u0, m0 = kurve[i]
        u1, m1 = kurve[i + 1]
        if u0 <= umsatz_meur <= u1:
            t = (umsatz_meur - u0) / (u1 - u0)
            return m0 + t * (m1 - m0)
    return kurve[-1][1]


def sozialplan_kosten_pro_person(entlassungen, anfangsbestand):
    if anfangsbestand <= 0 or entlassungen <= 0:
        return 0
    pct = entlassungen / anfangsbestand * 100
    if pct < 10:
        return 0
    if pct < 30:
        return 15
    if pct < 50:
        return 20
    return 25


class CalibrationEngine:
    MIN_PERIODS_FOR_AUTOFIT = 4
    MIN_PAIRS_POT = 3
    MIN_PAIRS_TECH = 3
    MIN_PAIRS_MVA = 2
    MIN_PAIRS_KZ = 2
    MIN_PERIODS_NACHARBEIT = 2
    MIN_PERIODS_GEHALTER = 2
    MIN_PERIODS_FIX = 2
    MIN_PERIODS_AKTIENKURS = 6
    FULL_CONF_OBS = 8
    BASIS_UMWELTINDEX = 91.5
    BASIS_FERTIGUNG = 40000

    DEFAULT_PARAMS = {
        "price_elasticity": -1.6,
        "werbung_exponent": 0.1,
        "tech_exponent": 0.3,
        "vertrieb_exponent": 0.15,
        "bekanntheit_exponent": 0.08,
        "kz_exponent": 0.1,
        "tech_per_meur": 1.56,
        "kz_price_sensitivity": -2.28,
        "kz_lager_penalty": -0.05,
        "kz_umwelt_bonus": 0.3,
        "kz_lieferunfaehig_penalty": -5.0,
        "preis_relativ_exp": -0.85,
        "lieferunfaehig_umverteilung": 0.80,
        "spillover_sweetspot": 0.0,
        "mva_delta_faktor": 2.44,
        "betriebsstoff_stueck": 50,
        "loehne_steigerung": 1.03,
        "transport_stueck": 25,
        "grossabnehmer_preis": 2700,
        "fixkosten_basis": 5.0,
        "einkauf_staffel": [
            [42000, 650], [60000, 550], [80000, 450], [999999, 400]
        ],
        "steuersatz": 0.40,
        "nacharbeit_basis_pct": 3.5,
        "nacharbeit_tech_factor": 0.02,
        "lagerkosten_fertig_stueck": 100,
        "lagerkosten_einsatz_stueck": 50,
        "ueberstunden_sprungfix": 2.5,
        "ueberstunden_lohnzuschlag": 0.25,
        "ueberstunden_max_pct": 0.10,
        "grundproduktivitaet": 50,
        "fehlzeiten_pct": 0.053,
        "ration_index_per_meur": 0.05,
        "wertanalyse_index": 1.0,
        "fe_personal_basis": 30,
        "tech_per_fe_ma": 0.1,
        "basiszins": 8.0,
        "gebaeude_abschreibung": 1.0,
        "verwaltung_instandhaltung": 1.0,
        "lohn_fertigung_teur": 31,
        "gehalt_vertrieb": 41,
        "gehalt_fe": 45,
        "gehalt_admin": 29,
        "gehalt_einkauf": 31,
        "sozialkosten_faktor": 1.48,
        "einstellungskosten_teur": 5,
        "entlassungskosten_teur": 10,
        "fluktuation_rate": 0.058,
        "transport_m2_stueck": 75,
        "wechselkurs": 0.125,
        "markt2_offen": False,
        "markt2_start_potenzial": 12000,
        "aktienkurs_w_ek": 0.25,
        "aktienkurs_w_netto": 0.20,
        "aktienkurs_w_umsatzrendite": 0.10,
        "aktienkurs_w_bekanntheit": 0.05,
        "aktienkurs_w_kz": 0.10,
        "aktienkurs_w_umsatz": 0.05,
        "aktienkurs_w_dividende": 0.05,
        "aktienkurs_w_fkquote": -0.10,
        "aktienkurs_w_techqual": 0.05,
        "aktienkurs_w_mva": 0.05,
    }

    def __init__(self, params_file="params_v6.json"):
        self.params_file = params_file
        self.params = self._load_params()

    def _load_params(self):
        if os.path.exists(self.params_file):
            with open(self.params_file) as f:
                stored = json.load(f)
            merged = deepcopy(self.DEFAULT_PARAMS)
            merged.update(stored)
            return merged
        return deepcopy(self.DEFAULT_PARAMS)

    def save_params(self):
        with open(self.params_file, "w") as f:
            json.dump(self.params, f, indent=2, ensure_ascii=False)

    def _confidence_weight(self, n_obs, min_obs, full_obs=None):
        if full_obs is None:
            full_obs = self.FULL_CONF_OBS
        if n_obs < min_obs:
            return 0.0
        if n_obs >= full_obs:
            return 1.0
        return (n_obs - min_obs + 1) / max(1, (full_obs - min_obs + 1))

    def _blend_with_default(self, key, fitted_value, n_obs, min_obs, digits=4, full_obs=None):
        """Stabilisiert Fits bei kleiner Datenbasis Richtung Handbuch-Defaults."""
        w = self._confidence_weight(n_obs, min_obs, full_obs=full_obs)
        default_val = self.DEFAULT_PARAMS.get(key, self.params.get(key, fitted_value))
        blended = default_val + w * (fitted_value - default_val)
        return round(blended, digits)

    def _reset_sparse_params(self, period_count):
        """Schuetzt bei kleiner Historie vor ausufernden, instabilen Parametern."""
        if period_count >= self.MIN_PERIODS_FOR_AUTOFIT:
            return
        clamp_ranges = {
            "price_elasticity": (-4.0, -0.2),
            "werbung_exponent": (0.0, 0.8),
            "tech_exponent": (0.0, 1.2),
            "tech_per_meur": (0.2, 4.0),
            "tech_per_fe_ma": (0.0, 0.3),
            "kz_price_sensitivity": (-5.0, 0.0),
        }
        for key, (lo, hi) in clamp_ranges.items():
            val = self.params.get(key)
            if isinstance(val, (int, float)):
                self.params[key] = max(lo, min(hi, float(val)))

    def calibrate(self, history):
        periods = history.all_periods()
        if len(periods) < 2:
            print("  Kalibrierung braucht min. 2 Perioden.")
            return self.params

        self._reset_sparse_params(len(periods))

        print(f"\n--- Kalibrierung mit {len(periods)} Perioden ---")
        self._fit_einkauf(history, periods)
        self._fit_gehaelter(history, periods)
        self._fit_cross_section_attractiveness(history, periods)
        self._fit_tech_per_meur(history, periods)
        self._fit_mva(history, periods)
        self._fit_kz(history, periods)
        self._fit_nacharbeit(history, periods)
        self._fit_sonstiger_aufwand(history, periods)
        self._fit_aktienkurs(history, periods)
        self._autofit_relationships(history, periods)
        self.save_params()
        print("  Parameter gespeichert.\n")
        return self.params

    def _autofit_relationships(self, history, periods):
        """Testet fuer Schlussel-Zusammenhaenge automatisch den besten Modelltyp.
        Nutzt Absolutwerte aller Perioden (nicht nur Uebergaenge)."""
        self.model_fits = {}
        all_data = [history.get_period(p) for p in periods]
        all_data = [d for d in all_data if d]
        if len(all_data) < self.MIN_PERIODS_FOR_AUTOFIT:
            print(f"\n  Auto-Fit: braucht min. {self.MIN_PERIODS_FOR_AUTOFIT} Perioden (haben {len(all_data)})")
            return

        relationships = [
            ("Preis->Absatz", "preis", "pot_absatz"),
            ("Werbung->Absatz", "werbung", "pot_absatz"),
            ("Tech->Absatz", "tech_index", "pot_absatz"),
            ("Preis->KZ", "preis", "kz_index"),
            ("Werbung->Umsatz", "werbung", "umsatz_gesamt"),
            ("F&E->Tech", "fe_invest_gen1", "tech_index"),
            ("Fertigung->Umsatz", "fertigungsmenge_tats", "umsatz_gesamt"),
            ("Fertigung->EBIT", "fertigungsmenge_tats", "ebit"),
            ("Personal->Kosten", "personal_summe", "personalaufwand"),
            ("Preis->Umsatz", "preis", "umsatz_gesamt"),
        ]

        print(f"\n  Auto-Fit Zusammenhangsanalyse ({len(all_data)} Perioden):")
        for label, x_key, y_key in relationships:
            x_vals, y_vals = [], []
            for d in all_data:
                xv = d.get(x_key)
                yv = d.get(y_key)
                if xv is not None and yv is not None:
                    x_vals.append(float(xv))
                    y_vals.append(float(yv))
            if len(x_vals) >= 2:
                typ, params, r2, fn = _fit_all_models(x_vals, y_vals)
                self.model_fits[label] = {"typ": typ, "params": params, "r2": r2, "fn": fn}
                tag = f"R²={r2:.4f}" if r2 > 0 else "n/a"
                print(f"    {label:22s}: {typ:13s} {_model_label(typ, params):45s} ({tag})")
            else:
                print(f"    {label:22s}: zu wenig Daten ({len(x_vals)} Punkte)")

    def _fit_pot_absatz(self, history, periods):
        pairs = []
        for i in range(1, len(periods)):
            prev = history.get_period(periods[i - 1])
            curr = history.get_period(periods[i])
            if not all(k in curr for k in ["pot_absatz", "preis", "werbung", "tech_index"]):
                continue
            if not all(k in prev for k in ["pot_absatz", "preis", "werbung", "tech_index"]):
                continue
            pairs.append((prev, curr))
        if len(pairs) < self.MIN_PAIRS_POT:
            print(f"  Pot.Absatz: zu wenig Uebergaenge ({len(pairs)}), behalte stabile Basiswerte.")
            return

        def residuals(x):
            p_el, w_exp, t_exp = x
            errs = []
            for prev, curr in pairs:
                predicted = prev["pot_absatz"] \
                    * (curr["preis"] / max(prev["preis"], 1)) ** p_el \
                    * (curr["werbung"] / max(prev["werbung"], 0.1)) ** w_exp \
                    * (curr["tech_index"] / max(prev["tech_index"], 1)) ** t_exp
                errs.append(predicted - curr["pot_absatz"])
            return errs

        x0 = [self.params["price_elasticity"], self.params["werbung_exponent"], self.params["tech_exponent"]]
        try:
            res = least_squares(residuals, x0, bounds=([-10, 0, 0], [0, 2, 5]))
            n_pairs = len(pairs)
            self.params["price_elasticity"] = self._blend_with_default(
                "price_elasticity", res.x[0], n_pairs, self.MIN_PAIRS_POT
            )
            self.params["werbung_exponent"] = self._blend_with_default(
                "werbung_exponent", res.x[1], n_pairs, self.MIN_PAIRS_POT
            )
            self.params["tech_exponent"] = self._blend_with_default(
                "tech_exponent", res.x[2], n_pairs, self.MIN_PAIRS_POT
            )
            print(
                "  Pot.Absatz:"
                f" p_el={self.params['price_elasticity']:.4f},"
                f" w_exp={self.params['werbung_exponent']:.4f},"
                f" t_exp={self.params['tech_exponent']:.4f}"
                f" (raw cost={res.cost:.1f}, n={n_pairs})"
            )
        except (ValueError, RuntimeError, OverflowError, FloatingPointError) as e:
            print(f"  Pot.Absatz-Fit fehlgeschlagen: {e}")

    def _fit_cross_section_attractiveness(self, history, periods):
        """Cross-sectional Nakanishi-Cooper Fit über alle Teams pro Periode.
        Extrahiert echte Elastizitäten bereinigt um Branchen-Mittelwert-Effekte.
        Fallback auf _fit_pot_absatz bei zu wenig Daten."""
        Y_data = []
        X_data = []  # [price_log, werbung_log, tech_log, vertrieb_log, bekanntheit_log, kz_log]

        for p in periods:
            pdata = history.get_period(p)
            if not pdata:
                continue
            alle = pdata.get("alle_unternehmen", {})
            if not alle:
                continue

            # Extrahiere Teamdaten, filtere ungültige
            teams = []
            for team_data in alle.values():
                absatz = team_data.get("pot_absatz", 0)
                preis = team_data.get("preis", 0)
                werbung = team_data.get("werbung", 0)
                tech = team_data.get("tech_index", 0)
                vertrieb = team_data.get("vertrieb_ma", 0)
                bekanntheit = team_data.get("bekanntheit", 0)
                kz = max(20.0, float(team_data.get("kz_index", 60.0)))
                if preis <= 0 or absatz <= 0:
                    continue
                werbung = max(werbung, 0.01)
                tech = max(tech, 0.01)
                vertrieb = max(vertrieb, 1.0)
                bekanntheit = max(bekanntheit, 1.0)
                teams.append((absatz, preis, werbung, tech, vertrieb, bekanntheit, kz))

            if len(teams) < 2:
                continue

            # Geometrische Mittelwerte (log-Mittel)
            log_absatz = [math.log(t[0]) for t in teams]
            log_preis   = [math.log(t[1]) for t in teams]
            log_werbung = [math.log(t[2]) for t in teams]
            log_tech    = [math.log(t[3]) for t in teams]
            log_vertrieb = [math.log(t[4]) for t in teams]
            log_bekanntheit = [math.log(t[5]) for t in teams]
            log_kz      = [math.log(t[6]) for t in teams]

            geom_absatz    = sum(log_absatz)    / len(teams)
            geom_preis     = sum(log_preis)     / len(teams)
            geom_werbung   = sum(log_werbung)   / len(teams)
            geom_tech      = sum(log_tech)      / len(teams)
            geom_vertrieb  = sum(log_vertrieb)  / len(teams)
            geom_bekanntheit = sum(log_bekanntheit) / len(teams)
            geom_kz        = sum(log_kz)        / len(teams)

            for i, t in enumerate(teams):
                Y_data.append(log_absatz[i] - geom_absatz)
                X_data.append([
                    log_preis[i]       - geom_preis,
                    log_werbung[i]     - geom_werbung,
                    log_tech[i]        - geom_tech,
                    log_vertrieb[i]    - geom_vertrieb,
                    log_bekanntheit[i] - geom_bekanntheit,
                    log_kz[i]          - geom_kz,
                ])

        if len(Y_data) < 6:
            print(
                f"  Cross-Section: zu wenig Beobachtungen ({len(Y_data)}), "
                f"Fallback auf Zeitreihen-Fit."
            )
            return self._fit_pot_absatz(history, periods)

        Y = np.array(Y_data)
        X = np.array(X_data)  # shape (n, 6)

        def residuals(params):
            pred = (params[0] * X[:, 0]
                    + params[1] * X[:, 1]
                    + params[2] * X[:, 2]
                    + params[3] * X[:, 3]
                    + params[4] * X[:, 4]
                    + params[5] * X[:, 5])
            return pred - Y

        x0 = [
            self.params["price_elasticity"],
            self.params["werbung_exponent"],
            self.params["tech_exponent"],
            self.params["vertrieb_exponent"],
            self.params["bekanntheit_exponent"],
            self.params.get("kz_exponent", 0.1),
        ]
        bounds_lo = [-5.0, 0.0, 0.0, 0.0, 0.0, 0.0]
        bounds_hi = [ 0.0, 2.0, 2.0, 2.0, 2.0, 2.0]

        try:
            res = least_squares(residuals, x0, bounds=(bounds_lo, bounds_hi))
            n_obs = len(Y_data)
            self.params["price_elasticity"] = self._blend_with_default(
                "price_elasticity", res.x[0], n_obs, 6, full_obs=24
            )
            self.params["werbung_exponent"] = self._blend_with_default(
                "werbung_exponent", res.x[1], n_obs, 6, full_obs=24
            )
            self.params["tech_exponent"] = self._blend_with_default(
                "tech_exponent", res.x[2], n_obs, 6, full_obs=24
            )
            self.params["vertrieb_exponent"] = self._blend_with_default(
                "vertrieb_exponent", res.x[3], n_obs, 6, full_obs=24
            )
            self.params["bekanntheit_exponent"] = self._blend_with_default(
                "bekanntheit_exponent", res.x[4], n_obs, 6, full_obs=24
            )
            self.params["kz_exponent"] = self._blend_with_default(
                "kz_exponent", res.x[5], n_obs, 6, full_obs=24
            )
            print(
                f"  Cross-Section Fit (n={n_obs}):"
                f" p_el={self.params['price_elasticity']:.4f},"
                f" w={self.params['werbung_exponent']:.4f},"
                f" t={self.params['tech_exponent']:.4f},"
                f" v={self.params['vertrieb_exponent']:.4f},"
                f" b={self.params['bekanntheit_exponent']:.4f},"
                f" kz={self.params['kz_exponent']:.4f}"
                f" (cost={res.cost:.2f})"
            )
        except (ValueError, RuntimeError, OverflowError, FloatingPointError) as e:
            print(f"  Cross-Section Fit fehlgeschlagen: {e}, Fallback auf Zeitreihen-Fit.")
            self._fit_pot_absatz(history, periods)

    def _fit_tech_per_meur(self, history, periods):
        """Fittet tech_per_meur und tech_per_fe_ma gemeinsam."""
        data_points = []
        for i in range(1, len(periods)):
            prev = history.get_period(periods[i - 1])
            curr = history.get_period(periods[i])
            if "tech_index" not in curr or "tech_index" not in prev:
                continue
            delta_tech = curr["tech_index"] - prev["tech_index"]
            fe_invest = curr.get("fe_invest_gen1") or curr.get("decisions", {}).get("fe_invest_gen1") or 0
            fe_personal = curr.get("personal_fe", prev.get("personal_fe", 34))
            data_points.append((delta_tech, fe_invest, fe_personal))

        if len(data_points) < self.MIN_PAIRS_TECH:
            print(f"  Tech-Fit: zu wenig Uebergaenge ({len(data_points)}), behalte stabile Basiswerte.")
            return

        fe_basis = self.params.get("fe_personal_basis", 30)

        def residuals(x):
            tech_meur, tech_ma = x
            return [dt - (tech_meur * fi + max(0, fp - fe_basis) * tech_ma)
                    for dt, fi, fp in data_points]

        x0 = [self.params["tech_per_meur"], self.params.get("tech_per_fe_ma", 0.1)]
        try:
            res = least_squares(residuals, x0, bounds=([0, 0], [10, 2]))
            n_obs = len(data_points)
            self.params["tech_per_meur"] = self._blend_with_default(
                "tech_per_meur", res.x[0], n_obs, self.MIN_PAIRS_TECH
            )
            self.params["tech_per_fe_ma"] = self._blend_with_default(
                "tech_per_fe_ma", res.x[1], n_obs, self.MIN_PAIRS_TECH
            )
            print(
                "  Tech:"
                f" per_MEUR={self.params['tech_per_meur']:.4f},"
                f" per_FE_MA={self.params['tech_per_fe_ma']:.4f}"
                f" (raw cost={res.cost:.1f}, n={n_obs})"
            )
        except (ValueError, RuntimeError, OverflowError, FloatingPointError) as e:
            print(f"  Tech-Fit fehlgeschlagen: {e}")

    def _fit_mva(self, history, periods):
        vals = []
        for i in range(1, len(periods)):
            prev = history.get_period(periods[i - 1])
            curr = history.get_period(periods[i])
            if "mva" in curr and "mva" in prev and "nopat" in curr and curr["nopat"] != 0:
                vals.append((curr["mva"] - prev["mva"]) / curr["nopat"])
        if len(vals) >= self.MIN_PAIRS_MVA:
            avg = sum(vals) / len(vals)
            self.params["mva_delta_faktor"] = self._blend_with_default(
                "mva_delta_faktor", avg, len(vals), self.MIN_PAIRS_MVA, full_obs=4
            )
            print(f"  MVA-Faktor: {self.params['mva_delta_faktor']:.4f} (n={len(vals)})")
        else:
            print(f"  MVA-Faktor: zu wenig Uebergaenge ({len(vals)}), unveraendert.")

    def _fit_kz(self, history, periods):
        vals = []
        for i in range(1, len(periods)):
            prev = history.get_period(periods[i - 1])
            curr = history.get_period(periods[i])
            if all(k in curr for k in ["kz_index", "preis"]) and all(k in prev for k in ["kz_index", "preis"]):
                dp = curr["preis"] - prev["preis"]
                dk = curr["kz_index"] - prev["kz_index"]
                if dp != 0:
                    vals.append(dk / dp * 100)
        if len(vals) >= self.MIN_PAIRS_KZ:
            avg = sum(vals) / len(vals)
            self.params["kz_price_sensitivity"] = self._blend_with_default(
                "kz_price_sensitivity", avg, len(vals), self.MIN_PAIRS_KZ
            )
            print(f"  KZ-Sensitivität: {self.params['kz_price_sensitivity']:.4f} (n={len(vals)})")
        else:
            print(f"  KZ-Sensitivität: zu wenig Uebergaenge ({len(vals)}), unveraendert.")

    def _fit_nacharbeit(self, history, periods):
        """Schaetzt Nacharbeit-% aus der Differenz zwischen realem und modelliertem
        Sonstigem Aufwand (nach Abzug von Umweltstrafe und Lagerkosten)."""
        vals = []
        for p_key in periods:
            d = history.get_period(p_key)
            if not d or "sonstiger_aufwand" not in d:
                continue
            real_sa = d["sonstiger_aufwand"]
            umw_idx = d.get("umweltindex_anlagen", self.BASIS_UMWELTINDEX)
            lager = d.get("lager_fertig", 0)
            geschaetzte_strafe = umwelt_strafe(umw_idx)
            lagerkosten = lager * self.params["lagerkosten_fertig_stueck"] / 1e6
            rest = real_sa - geschaetzte_strafe - lagerkosten - self.params.get("gebaeude_abschreibung", 1.0) - self.params.get("verwaltung_instandhaltung", 1.0)
            fert = d.get("fertigungsmenge_tats", self.BASIS_FERTIGUNG)
            einkauf = 550
            material_meur = fert * einkauf / 1e6
            betriebs_meur = fert * self.params["betriebsstoff_stueck"] / 1e6
            basis_kosten = material_meur + betriebs_meur
            if basis_kosten > 0 and rest > 0:
                pct = rest / basis_kosten * 100
                vals.append(min(pct, 15.0))
        if len(vals) >= max(self.MIN_PERIODS_NACHARBEIT, 4):
            avg = sum(vals) / len(vals)
            self.params["nacharbeit_basis_pct"] = self._blend_with_default(
                "nacharbeit_basis_pct", avg, len(vals), self.MIN_PERIODS_NACHARBEIT, digits=2, full_obs=4
            )
            print(f"  Nacharbeit: {self.params['nacharbeit_basis_pct']:.2f}% (n={len(vals)})")
        else:
            self.params["nacharbeit_basis_pct"] = 4.9
            print(
                f"  Nacharbeit: zu wenig belastbare Punkte ({len(vals)}), "
                f"fixiert auf {self.params['nacharbeit_basis_pct']}%."
            )

    def _fit_einkauf(self, history, periods):
        """Kalibriert Einkaufsstaffel aus realen var_materialkosten/Stueck."""
        for p_key in periods:
            d = history.get_period(p_key)
            if d and "var_materialkosten" in d and "fertigungsmenge_tats" in d:
                real_preis = d["var_materialkosten"]
                fert = d["fertigungsmenge_tats"]
                if real_preis > 0 and fert > 0:
                    staffel = self.params["einkauf_staffel"]
                    for tier in staffel:
                        if fert <= tier[0]:
                            if abs(tier[1] - real_preis) > 5:
                                tier[1] = round(real_preis, 0)
                                print(f"  Einkauf: {fert:.0f} Stk -> {real_preis:.0f} EUR (Staffel angepasst)")
                            break

    def _fit_gehaelter(self, history, periods):
        """Kalibriert den Sozialkosten-Faktor aus echtem Personalaufwand vs. Basis-Gehaeltern.
        Zieht geschaetzte Personalwechselkosten (Fluktuation, Ein-/Entlassungen) ab,
        damit der Faktor nur Lohnnebenkosten erfasst (keine Doppelzaehlung)."""
        faktor_list = []
        flukt_rate = self.params.get("fluktuation_rate", 0.058)
        for p_key in periods:
            d = history.get_period(p_key)
            if not d or "personalaufwand" not in d:
                continue
            real_pa = d["personalaufwand"]
            fert = d.get("personal_fertigung", 852)
            vertrieb = d.get("personal_vertrieb", 100)
            fe = d.get("personal_fe", 34)
            einkauf = d.get("personal_einkauf", 20)
            total = d.get("personal_summe", fert + vertrieb + fe + einkauf + 200)
            admin = max(0, total - fert - vertrieb - fe - einkauf)

            flukt = round(total * flukt_rate)
            einstell_k = self.params.get("einstellungskosten_teur", 5)
            entlass_k = self.params.get("entlassungskosten_teur", 10)
            pers_aend_fert = d.get("decisions", {}).get("personal_aenderung_fert", 0)
            einst_gesamt = max(0, pers_aend_fert) + flukt
            entl_gesamt = max(0, -pers_aend_fert)
            wechsel_kosten = (einst_gesamt * einstell_k + entl_gesamt * entlass_k) / 1e3

            pa_ohne_wechsel = real_pa - wechsel_kosten

            lohn_f = self.params.get("lohn_fertigung_teur", 31)
            geh_v = self.params.get("gehalt_vertrieb", 41)
            geh_fe = self.params.get("gehalt_fe", 45)
            geh_a = self.params.get("gehalt_admin", 29)
            geh_e = self.params.get("gehalt_einkauf", 31)
            basis_pa = (fert * lohn_f + vertrieb * geh_v + fe * geh_fe
                        + admin * geh_a + einkauf * geh_e) / 1e3
            if basis_pa > 0 and pa_ohne_wechsel > 0:
                faktor_list.append(pa_ohne_wechsel / basis_pa)
        if len(faktor_list) >= self.MIN_PERIODS_GEHALTER:
            avg_faktor = sum(faktor_list) / len(faktor_list)
            self.params["sozialkosten_faktor"] = self._blend_with_default(
                "sozialkosten_faktor", avg_faktor, len(faktor_list), self.MIN_PERIODS_GEHALTER, full_obs=4
            )
            print(f"  Sozialkosten-Faktor: {self.params['sozialkosten_faktor']:.4f} (n={len(faktor_list)})")
        else:
            print(f"  Sozialkosten-Faktor: zu wenig Perioden ({len(faktor_list)}), unveraendert.")

    def _fit_sonstiger_aufwand(self, history, periods):
        """Kalibriert fixkosten_basis: restliche Kosten nach Abzug modellierter Bestandteile."""
        vals = []
        for p_key in periods:
            d = history.get_period(p_key)
            if not d:
                continue
            real_ebit = d.get("ebit", d.get("betriebsergebnis"))
            real_umsatz = d.get("umsatz_gesamt")
            real_bv = d.get("bestandsveraenderung_guv", 0)
            real_personal = d.get("personalaufwand", 0)
            real_material = d.get("materialaufwand", 0)
            real_abschr = d.get("abschreibungen", 0)
            real_sonst = d.get("sonstiger_aufwand", 0)
            if real_ebit is not None and real_umsatz is not None:
                fert = d.get("fertigungsmenge_tats", self.BASIS_FERTIGUNG)
                werbung = d.get("werbung", 6)
                fe_inv = d.get("fe_invest_gen1", 0) + d.get("fe_invest_gen2", 0)
                umw_idx = d.get("umweltindex_anlagen", self.BASIS_UMWELTINDEX)
                betriebs = fert * self.params["betriebsstoff_stueck"] / 1e6
                transport = d.get("absatz_gesamt", fert) * self.params["transport_stueck"] / 1e6
                umw_strafe = umwelt_strafe(umw_idx)
                lager = d.get("lager_fertig", 0)
                lagerk = lager * self.params["lagerkosten_fertig_stueck"] / 1e6
                modellierte_sonst = betriebs + transport + werbung + fe_inv + umw_strafe + lagerk + 2.0
                rest = real_sonst - modellierte_sonst
                if rest > 0:
                    vals.append(rest)
        if len(vals) >= self.MIN_PERIODS_FIX:
            avg = sum(vals) / len(vals)
            self.params["fixkosten_basis"] = self._blend_with_default(
                "fixkosten_basis", avg, len(vals), self.MIN_PERIODS_FIX, digits=2, full_obs=4
            )
            print(f"  Fixkosten-Basis: {self.params['fixkosten_basis']:.2f} MEUR (n={len(vals)})")
        else:
            print(f"  Fixkosten-Basis: zu wenig Perioden ({len(vals)}), unveraendert.")

    def _fit_aktienkurs(self, history, periods):
        """Fittet Aktienkurs-Gewichte mit Least Squares."""
        data_points = []
        for p_key in periods:
            d = history.get_period(p_key)
            if d and "aktienkurs" in d and "eigenkapital" in d and "mva" in d:
                data_points.append(d)
        if len(data_points) < self.MIN_PERIODS_AKTIENKURS:
            print(f"  Aktienkurs-Fit: zu wenig Perioden ({len(data_points)}), behalte stabile Basisgewichte.")
            return

        param_keys = [
            "aktienkurs_w_ek", "aktienkurs_w_netto", "aktienkurs_w_umsatzrendite",
            "aktienkurs_w_bekanntheit", "aktienkurs_w_kz", "aktienkurs_w_umsatz",
            "aktienkurs_w_dividende", "aktienkurs_w_fkquote", "aktienkurs_w_techqual", "aktienkurs_w_mva",
        ]

        def predict(weights, dp):
            ek = dp.get("eigenkapital", 30)
            mva_val = dp.get("mva", 60)
            netto = dp.get("periodenueberschuss", 5)
            ur = dp.get("umsatzrendite", 5)
            bek = dp.get("bekanntheit", 50)
            kz = dp.get("kz_index", 60)
            ums = dp.get("umsatz_gesamt", 130)
            fkq = dp.get("fremdkapitalquote", 50)
            tech = dp.get("tech_index", 100)
            score = (
                weights[0] * ek + weights[1] * netto * 5 + weights[2] * ur * 3
                + weights[3] * bek * 0.5 + weights[4] * kz * 0.5
                + weights[5] * ums * 0.3 + weights[6] * 0
                + weights[7] * fkq * 0.3 + weights[8] * tech * 0.3
                + weights[9] * mva_val * 0.5
            )
            return max(10, (ek + mva_val) * 2 * 0.5 + score * 0.5)

        def residuals(weights):
            return [predict(weights, dp) - dp["aktienkurs"] for dp in data_points]

        if len(data_points) <= len(param_keys):
            print(
                "  Aktienkurs-Fit: unteridentifiziert "
                f"({len(data_points)} Punkte fuer {len(param_keys)} Gewichte), uebersprungen."
            )
            return

        x0 = [self.params[k] for k in param_keys]
        try:
            res = least_squares(residuals, x0, bounds=(-2, 2))
            for i, k in enumerate(param_keys):
                self.params[k] = round(res.x[i], 4)
            print(f"  Aktienkurs-Fit: cost={res.cost:.1f} (aus {len(data_points)} Datenpunkten)")
        except (ValueError, RuntimeError, OverflowError, FloatingPointError) as e:
            print(f"  Aktienkurs-Fit fehlgeschlagen: {e}")


# ---------------------------------------------------------------------------
# Core Simulator V5.1
# ---------------------------------------------------------------------------

class TOPSIM_EagleEye_V5:
    N_TEAMS = 6
    NEWS_FILE = "news_v6.json"
    BASIS_KUMUL_FERTIGUNG = 40000
    BASIS_UMWELTINDEX = 91.5
    BASIS_LAGER_FERTIG = 9000
    ANLAGEN_KAP_A = 14000
    ANLAGEN_KAP_B = 30000
    MAX_GROSSABNEHMER = 9000

    BEKANNTE_NEWS = {
        0: {
            "lohn_einkauf": 30, "lohn_verwaltung": 28, "lohn_fertigung": 30,
            "lohn_fe": 44, "lohn_vertrieb": 40,
            "betriebsstoff": 50, "basiszins": 8.0,
            "einstellungskosten": 12.5, "entlassungskosten": 10,
            "transport_m1": 25,
        },
        1: {
            "lohn_einkauf": 31, "lohn_verwaltung": 29, "lohn_fertigung": 31,
            "lohn_fe": 45, "lohn_vertrieb": 41,
            "betriebsstoff": 50, "basiszins": 8.0,
            "einstellungskosten": 5, "entlassungskosten": 10,
            "transport_m1": 25,
        },
        2: {
            "lohn_einkauf": 32, "lohn_verwaltung": 30, "lohn_fertigung": 32,
            "lohn_fe": 46, "lohn_vertrieb": 42,
            "betriebsstoff": 65, "basiszins": 9.0,
            "einstellungskosten": 5, "entlassungskosten": 10,
            "transport_m1": 25, "transport_m2": 75,
            "wechselkurs": 0.125, "markt2_offen": True,
        },
    }

    def __init__(self, reports_dir="reports", state_file="topsim_state_v6.json",
                 history_file="history_v6.json", params_file="params_v6.json",
                 unternehmen_nr=2):
        self.reports_dir = reports_dir
        self.state_file = state_file
        self.u_nr = unternehmen_nr
        self.history = HistoryDB(history_file)
        self.calibration = CalibrationEngine(params_file)
        self.decisions = {}
        self._scenario_nr = 0
        self.news = self._load_news()

        imported = self._ingest_all_reports()
        if imported == 0 and self.history.period_count() >= 2:
            # Auch ohne neue Dateien periodisch stabilisieren (Sparse-Reset etc.)
            self.calibration.calibrate(self.history)
        self._load_state()

        p = self.calibration.params
        print(f"\n{'='*65}")
        print(f"  TOPSIM Eagle Eye V6.1 – Periode {self.state['periode']}")
        print(f"  History: {self.history.period_count()} Perioden geladen")
        print(f"  News: {len(self.news)} Perioden hinterlegt")
        print(f"  Params: p_el={p['price_elasticity']:.3f} | w_exp={p['werbung_exponent']:.3f} | t_exp={p['tech_exponent']:.3f}")
        print(f"  Tech/MEUR={p['tech_per_meur']:.3f} | MVA-F={p['mva_delta_faktor']:.3f}")
        print(f"{'='*65}\n")

    # ---- News Management ----

    def _load_news(self):
        news = deepcopy(self.BEKANNTE_NEWS)
        if os.path.exists(self.NEWS_FILE):
            with open(self.NEWS_FILE) as f:
                stored = json.load(f)
            for k, v in stored.items():
                news[int(k)] = v
        return news

    def _save_news(self):
        with open(self.NEWS_FILE, "w") as f:
            json.dump({str(k): v for k, v in self.news.items()}, f, indent=2, ensure_ascii=False)

    def get_news(self, periode):
        return self.news.get(periode, {})

    def _is_markt2_open_for_period(self, periode):
        offen = bool(self.calibration.params.get("markt2_offen", False))
        for p_nr in sorted(self.news.keys()):
            if p_nr > periode:
                break
            news = self.news.get(p_nr, {})
            if "markt2_offen" in news:
                offen = bool(news["markt2_offen"])
        return offen

    def set_news(self, periode, news_dict):
        self.news[periode] = news_dict
        self._save_news()
        print(f"  News fuer Periode {periode} gespeichert ({len(news_dict)} Werte).")

    def apply_news(self, periode=None):
        """Wendet Wirtschaftsnachrichten fuer eine Periode auf params + state an."""
        if periode is None:
            periode = self.state["periode"] + 1
        news = self.get_news(periode)
        if not news:
            print(f"  Keine News fuer Periode {periode} hinterlegt.")
            return

        p = self.calibration.params
        s = self.state
        applied = []

        if "lohn_fertigung" in news:
            p["lohn_fertigung_teur"] = news["lohn_fertigung"]
            applied.append(f"Fertigung={news['lohn_fertigung']}k")
        if "lohn_vertrieb" in news:
            p["gehalt_vertrieb"] = news["lohn_vertrieb"]
            applied.append(f"Vertrieb={news['lohn_vertrieb']}k")
        if "lohn_fe" in news:
            p["gehalt_fe"] = news["lohn_fe"]
            applied.append(f"F&E={news['lohn_fe']}k")
        if "lohn_verwaltung" in news:
            p["gehalt_admin"] = news["lohn_verwaltung"]
            applied.append(f"Verwalt.={news['lohn_verwaltung']}k")
        if "lohn_einkauf" in news:
            p["gehalt_einkauf"] = news["lohn_einkauf"]
            applied.append(f"Einkauf={news['lohn_einkauf']}k")

        if "betriebsstoff" in news:
            p["betriebsstoff_stueck"] = news["betriebsstoff"]
            applied.append(f"Betriebsstoffe={news['betriebsstoff']} EUR/Stk")
        if "anlagen_kap_b" in news:
            p["anlagen_kap_b"] = news["anlagen_kap_b"]
            applied.append(f"AnlagenKapB={news['anlagen_kap_b']}")
        if "luftfracht_preis" in news:
            p["luftfracht_preis"] = news["luftfracht_preis"]
            applied.append(f"Luftfracht={news['luftfracht_preis']}")
        if "einkauf_staffel" in news and isinstance(news["einkauf_staffel"], str):
            try:
                staffel = [[int(k.strip()), int(v.strip())] for pair in news["einkauf_staffel"].split(",") for k, v in [pair.split(":")]]
                if staffel:
                    p["einkauf_staffel"] = staffel
                    applied.append("Einkaufsstaffel")
            except Exception:
                pass

        if "basiszins" in news:
            p["basiszins"] = news["basiszins"]
            applied.append(f"Basiszins={news['basiszins']}%")

        if "transport_m1" in news:
            p["transport_stueck"] = news["transport_m1"]
        if "transport_m2" in news:
            p["transport_m2_stueck"] = news["transport_m2"]
            applied.append(f"Transport M2={news['transport_m2']} EUR/Stk")

        if "einstellungskosten" in news:
            p["einstellungskosten_teur"] = news["einstellungskosten"]
            applied.append(f"Einstell.kosten={news['einstellungskosten']}k")

        if "entlassungskosten" in news:
            p["entlassungskosten_teur"] = news["entlassungskosten"]

        if "wechselkurs" in news:
            p["wechselkurs"] = news["wechselkurs"]
            applied.append(f"Wechselkurs={news['wechselkurs']}")

        if "markt2_offen" in news:
            p["markt2_offen"] = bool(news["markt2_offen"])
            s["markt2_offen"] = bool(news["markt2_offen"])
            applied.append(f"Markt2={'offen' if news['markt2_offen'] else 'geschlossen'}")

        if "bip_wachstum" in news:
            boost = 1 + news["bip_wachstum"] / 100.0
            s["pot_absatz_vor"] = s.get("pot_absatz_vor", 43000) * boost
            applied.append(f"BIP +{news['bip_wachstum']}%")

        self.calibration.save_params()
        self._save_state()
        print(f"\n  News P{periode} angewendet: {', '.join(applied)}")

    def print_news(self, periode=None):
        if periode is None:
            periode = self.state["periode"] + 1
        news = self.get_news(periode)
        if not news:
            print(f"  Keine News fuer Periode {periode}.")
            return
        print(f"\n  {'='*60}")
        print(f"  WIRTSCHAFTSNACHRICHTEN – Periode {periode}")
        print(f"  {'='*60}")
        labels = {
            "lohn_einkauf": ("Lohn Einkauf", "TEUR/MA"),
            "lohn_verwaltung": ("Lohn Verwaltung", "TEUR/MA"),
            "lohn_fertigung": ("Lohn Fertigung", "TEUR/MA"),
            "lohn_fe": ("Lohn F&E", "TEUR/MA"),
            "lohn_vertrieb": ("Lohn Vertrieb", "TEUR/MA"),
            "betriebsstoff": ("Betriebsstoffe Gen1", "EUR/Stk"),
            "basiszins": ("Zinsen Ueberziehungskr.", "%"),
            "transport_m1": ("Transport Markt 1", "EUR/Stk"),
            "transport_m2": ("Transport Markt 2", "EUR/Stk"),
            "einstellungskosten": ("Einstellungskosten", "TEUR/MA"),
            "entlassungskosten": ("Entlassungskosten", "TEUR/MA"),
            "anlagen_kap_b": ("Kapazität Typ B", "Stk"),
            "luftfracht_preis": ("Luftfracht Preis", "EUR/Stk"),
            "einkauf_staffel": ("Einkaufsstaffel", ""),
            "wechselkurs": ("Wechselkurs EUR/FCU", ""),
            "bip_wachstum": ("BIP-Wachstum", "%"),
            "markt2_offen": ("Markt 2 offen", ""),
        }
        for key, val in news.items():
            label, unit = labels.get(key, (key, ""))
            print(f"    {label:28s}: {val} {unit}")
        print(f"  {'='*60}")

    # ---- Data Ingestion ----

    def _ingest_all_reports(self, quiet=False):
        patterns = [os.path.join(self.reports_dir, "*.xls"), os.path.join(self.reports_dir, "*.xlsx")]
        files = []
        for pat in patterns:
            files.extend(globmod.glob(pat))
        if not files:
            if not quiet:
                print(f"  Keine Reports in {self.reports_dir}/ gefunden.")
            return 0
        new_count = 0
        for fpath in sorted(files):
            try:
                parser = ReportParser(fpath, self.u_nr)
                data = parser.parse()
                p_nr = data.get("periode", -1)
                if p_nr < 0:
                    continue
                existing = self.history.get_period(p_nr)
                f_mtime = os.path.getmtime(fpath)
                f_size = os.path.getsize(fpath)
                if (
                    existing
                    and existing.get("_source") == fpath
                    and existing.get("_source_mtime") == f_mtime
                    and existing.get("_source_size") == f_size
                ):
                    continue
                data["_source_mtime"] = f_mtime
                data["_source_size"] = f_size
                self.history.add_period(p_nr, data)
                new_count += 1
                if not quiet:
                    print(f"  Report P{p_nr} eingelesen: {os.path.basename(fpath)}")
            except Exception as e:
                if not quiet:
                    print(f"  Fehler beim Parsen von {fpath}: {e}")
        if new_count > 0:
            if not quiet:
                print(f"  {new_count} neue Report(s) importiert.")
            self.calibration.calibrate(self.history)
        elif not quiet:
            print(f"  Alle {len(files)} Reports bereits importiert.")
        return new_count

    def ingest_report(self, path):
        parser = ReportParser(path, self.u_nr)
        data = parser.parse()
        p_nr = data.get("periode", -1)
        data["_source_mtime"] = os.path.getmtime(path)
        data["_source_size"] = os.path.getsize(path)
        self.history.add_period(p_nr, data)
        print(f"  Report P{p_nr} importiert: {os.path.basename(path)}")
        self.calibration.calibrate(self.history)
        self._load_state()

    # ---- State Management ----

    def _load_state(self):
        latest_p = self.history.all_periods()
        if latest_p:
            last = self.history.get_period(latest_p[-1])
            self.state = self._state_from_report(last)
        else:
            self.state = self._default_state()
        self._save_state()

    def _state_from_report(self, r):
        personal_fert = r.get("personal_fert_end", r.get("personal_fertigung", 852))
        personal_vertrieb = r.get("personal_vertrieb_end", r.get("personal_vertrieb", 100))
        personal_fe = r.get("personal_fe_end", r.get("personal_fe", 34))
        return {
            "periode": r.get("periode", 0),
            "tech_index": r.get("tech_index", 100.0),
            "lager_fertig": r.get("lager_fertig", 0),
            "lager_wert": r.get("lager_wert", 0),
            "anlagevermoegen": r.get("anlagevermoegen", 42.25),
            "ueberziehung": r.get("ueberziehungskredit", 0),
            "eigenkapital": r.get("eigenkapital", 31.52),
            "mva": r.get("mva", 61.7),
            "kz_index": r.get("kz_index", 69.97),
            "bekanntheit": r.get("bekanntheit", 49.25),
            "personal_fert": personal_fert,
            "personal_vertrieb": personal_vertrieb,
            "personal_fe": personal_fe,
            "personal_einkauf": r.get("personal_einkauf", 20),
            "abschreibungen": r.get("abschreibungen", 7.75),
            "preis_vor": r.get("preis", 3000),
            "werbung_vor": r.get("werbung", 6.0),
            "pot_absatz_vor": r.get("pot_absatz", 43000),
            "anlagen_kapazitaet": r.get("anlagen_kapazitaet", 42000),
            "nopat_vor": r.get("nopat", 0),
            "wacc": r.get("wacc", 7.43),
            "rating": r.get("rating", "BB"),
            "umweltindex": r.get("umweltindex_anlagen", self.BASIS_UMWELTINDEX),
            "umsatzrendite_vor": r.get("umsatzrendite", 5.0),
            "fremdkapitalquote_vor": r.get("fremdkapitalquote", 63.67),
            "motivation": r.get("motivation", 50.0),
            "kumul_fertigung": r.get("kumul_fertigung", self.BASIS_KUMUL_FERTIGUNG),
            "produktivitaet_idx1": r.get("produktivitaet_index1", 1.0),
            "produktivitaet_idx2": r.get("produktivitaet_index2", 1.0),
            "nicht_gedeckt_vor": r.get("nicht_gedeckt", 0),
            "kumul_dividende": 0,
            "rationalisierung_index": 1.0,
            "forderungen_vor": r.get("forderungen", 25.8),
            "branche_avg_preis": r.get("branche_avg_preis", 3000),
            "branche_nicht_gedeckt": r.get("branche_nicht_gedeckt_summe", 0),
            "branche_pot_absatz": r.get("branche_pot_absatz_summe", 258000),
            "loehne_basis": r.get("loehne", 37.46),
            "personal_summe": r.get("personal_summe", 1213),
            "personal_admin": r.get("personal_admin", 208),
            "verlustvortrag": r.get("verlustvortrag", 0),
            "sonstiger_aufwand_real": r.get("sonstiger_aufwand", 19.46),
            "bestandsveraenderung_real": r.get("bestandsveraenderung_guv", 0),
            "markt2_offen": r.get("markt2_offen", False),
            "pot_absatz_m2_vor": r.get("pot_absatz_m2", 0),
            "preis_m2_vor": r.get("preis_m2", 4200),
            "werbung_m2_vor": r.get("werbung_m2", 0),
        }

    def _default_state(self):
        return {
            "periode": 0, "tech_index": 100.0, "lager_fertig": self.BASIS_LAGER_FERTIG,
            "lager_wert": 18.62, "anlagevermoegen": 42.25, "ueberziehung": 43.37,
            "eigenkapital": 31.52, "mva": 61.7, "kz_index": 69.97,
            "bekanntheit": 49.25, "personal_fert": 852, "personal_vertrieb": 100,
            "personal_fe": 34, "personal_einkauf": 20, "abschreibungen": 7.75, "preis_vor": 3000,
            "werbung_vor": 6.0, "pot_absatz_vor": 43000, "anlagen_kapazitaet": 42000,
            "nopat_vor": 8.35, "wacc": 7.43, "rating": "BB", "umweltindex": self.BASIS_UMWELTINDEX,
            "umsatzrendite_vor": 5.06, "fremdkapitalquote_vor": 63.67,
            "motivation": 50.78, "kumul_fertigung": self.BASIS_KUMUL_FERTIGUNG,
            "produktivitaet_idx1": 1.0, "produktivitaet_idx2": 1.0,
            "nicht_gedeckt_vor": 0, "kumul_dividende": 0,
            "rationalisierung_index": 1.0, "forderungen_vor": 25.8,
            "branche_avg_preis": 3000, "branche_nicht_gedeckt": 0,
            "branche_pot_absatz": 258000, "loehne_basis": 37.46,
            "personal_summe": 1205, "personal_admin": 200,
            "verlustvortrag": 0,
            "sonstiger_aufwand_real": 19.46, "bestandsveraenderung_real": 0,
            "markt2_offen": False,
            "pot_absatz_m2_vor": 0,
            "preis_m2_vor": 4200,
            "werbung_m2_vor": 0,
        }

    def _save_state(self):
        with open(self.state_file, "w") as f:
            json.dump(self.state, f, indent=4, ensure_ascii=False)

    # ---- Optimierer: Rechnerisch beste Entscheidungen ----

    DECISION_BOUNDS = {
        "preis_m1":              (2800, 4200),
        "werbung_m1":            (4.0,  18.0),
        "vertrieb_ma":           (70,   160),
        "fertigungsmenge":       (30000, 55000),
        "fe_invest_gen1":        (25.0, 90.0),
        "oeko_budget":           (0.0,  2.0),
        "ci_budget":             (0.0,  2.0),
        "rationalisierung":      (0.0,  5.0),
        "wertanalyse":           (0.0,  2.0),
        "neue_anlagen_a":        (0,    3),
        "neue_anlagen_b":        (0,    1),
        "personal_aenderung_fert": (-50, 100),
        "fe_personal_aenderung": (-15,  15),
        "grossabnehmer":         (0,    MAX_GROSSABNEHMER),
        "kredit":                (0.0,  20.0),
        "dividende":             (0.0,  5.0),
    }

    ZIEL_GEWICHTE = {
        "aktienkurs":       {"w": 0.30, "label": "Aktienkurs"},
        "ebit":             {"w": 0.20, "label": "EBIT"},
        "eigenkapital":     {"w": 0.15, "label": "Eigenkapital"},
        "mva":              {"w": 0.10, "label": "MVA"},
        "kz_index":         {"w": 0.10, "label": "KZ-Index"},
        "periodenueberschuss": {"w": 0.10, "label": "Per.Ueberschuss"},
        "tech_index":       {"w": 0.05, "label": "Tech-Index"},
        "marktanteil":      {"w": 0.28, "label": "Marktanteil"},
    }

    def predict_competitors(self, target_periode, mode="advanced"):
        """V6.2: MCI-Attraktions-Modell (mode='advanced') oder naiver Trend (mode='simple').
        Gibt ein Dictionary mit aggregierten Marktdaten zurueck oder None.
        """
        periods = self.history.all_periods()
        if len(periods) < 2:
            return None

        data_t = self.history.get_period(periods[-1])
        data_t1 = self.history.get_period(periods[-2])
        if not data_t or not data_t1:
            return None

        alle_t = data_t.get("alle_unternehmen", {})
        alle_t1 = data_t1.get("alle_unternehmen", {})
        if not alle_t or not alle_t1:
            return None

        own = f"U{self.u_nr}"

        if mode == "simple":
            # --- Naiver Trend-Modus (unveraendert V6.1) ---
            competitor_predictions = {}
            for u in [f"U{i}" for i in range(1, 7)]:
                if u == own:
                    continue
                comp_t = alle_t.get(u, {})
                comp_t1 = alle_t1.get(u, {})
                preis_t = comp_t.get("preis", 0)
                preis_t_1 = comp_t1.get("preis", 0)
                if preis_t > 0 and preis_t_1 > 0:
                    preis_trend = (preis_t - preis_t_1) * 0.45
                    absatz_t = comp_t.get("nicht_gedeckt", 0)
                    absatz_t_1 = comp_t1.get("nicht_gedeckt", 0)
                    absatz_trend = (absatz_t - absatz_t_1) * 0.45
                    pred_preis = max(2800.0, min(4500.0, preis_t + preis_trend))
                    pred_nicht_gedeckt = max(0.0, absatz_t + absatz_trend)
                else:
                    pred_preis = preis_t if preis_t > 0 else 3000.0
                    pred_nicht_gedeckt = comp_t.get("nicht_gedeckt", 0)
                menge_t = comp_t.get("tats_absatz_markt", 0) + comp_t.get("nicht_gedeckt", 0)
                menge_t1 = comp_t1.get("tats_absatz_markt", 0) + comp_t1.get("nicht_gedeckt", 0)
                predicted_menge = max(0, menge_t + (menge_t - menge_t1)) if (menge_t > 0 and menge_t1 > 0) else menge_t
                competitor_predictions[u] = {
                    "predicted_preis": pred_preis,
                    "predicted_menge": predicted_menge,
                    "predicted_nicht_gedeckt": pred_nicht_gedeckt,
                }
            if not competitor_predictions:
                return None
            preise = [v["predicted_preis"] for v in competitor_predictions.values()]
            predicted_branche_avg_preis = sum(preise) / len(preise) if preise else 3000
            predicted_branche_nicht_gedeckt = sum(
                v["predicted_nicht_gedeckt"] for v in competitor_predictions.values()
            )
            return {
                "predicted_branche_avg_preis": predicted_branche_avg_preis,
                "predicted_branche_avg_kz": data_t.get("branche_avg_kz", 60.0),
                "predicted_branche_nicht_gedeckt": predicted_branche_nicht_gedeckt,
                "competitor_details": competitor_predictions,
            }

        # --- MCI Advanced-Modus ---
        p = self.calibration.params
        team_preds = {}

        for u in [f"U{i}" for i in range(1, 7)]:
            comp_t = alle_t.get(u, {})
            comp_t1 = alle_t1.get(u, {})

            def _trend(key, default, lo, hi):
                v_t = comp_t.get(key, default)
                v_t1 = comp_t1.get(key, default)
                if v_t > 0 and v_t1 > 0:
                    return max(lo, min(hi, v_t + (v_t - v_t1) * 0.45))
                return max(lo, min(hi, v_t if v_t > 0 else default))

            pred_preis    = _trend("preis",         3000.0, 2200.0, 6200.0)
            pred_werbung  = _trend("werbung",          1.0,    0.1, 9999.0)
            pred_tech     = _trend("tech_index",      100.0,   80.0, 9999.0)
            pred_vertrieb = _trend("vertrieb_ma",     100.0,   60.0, 9999.0)
            pred_kz       = max(25.0, min(95.0, comp_t.get("kz_index", 60.0)))

            tats_t  = comp_t.get("tats_absatz_markt", 0)
            tats_t1 = comp_t1.get("tats_absatz_markt", 0)

            team_preds[u] = {
                "pred_preis": pred_preis,
                "pred_werbung": pred_werbung,
                "pred_tech": pred_tech,
                "pred_vertrieb": pred_vertrieb,
                "pred_kz": pred_kz,
                "tats_t": tats_t,
                "tats_t1": tats_t1,
            }

        # Branchenschnitt Preis fuer relativen Term
        branche_avg_preis_pred = sum(v["pred_preis"] for v in team_preds.values()) / max(1, len(team_preds))

        # Attraktivitaeten berechnen
        attractiveness = {}
        for u, tv in team_preds.items():
            try:
                A = (
                    (tv["pred_preis"]    ** p["price_elasticity"])
                    * (tv["pred_werbung"]  ** p["werbung_exponent"])
                    * (tv["pred_tech"]     ** p["tech_exponent"])
                    * (tv["pred_vertrieb"] ** p["vertrieb_exponent"])
                    * (tv["pred_kz"]       ** p.get("kz_exponent", 0.12))
                )
                attractiveness[u] = max(A, 1e-9)
            except (ZeroDivisionError, OverflowError, ValueError):
                attractiveness[u] = 1e-9

        sum_A = sum(attractiveness.values())
        total_market = data_t.get("branche_pot_absatz_summe", 258000)

        competitor_predictions = {}
        for u, tv in team_preds.items():
            if u == own:
                continue
            pred_pot_absatz = total_market * (attractiveness[u] / max(sum_A, 1e-9))
            comp_data = alle_t.get(u, {})
            cap_j = float(comp_data.get("tats_absatz_markt", 0)) * 1.1
            pred_nicht_gedeckt = max(0.0, pred_pot_absatz - cap_j)
            competitor_predictions[u] = {
                "predicted_preis": tv["pred_preis"],
                "predicted_menge": tv["tats_t"],
                "predicted_pot_absatz": pred_pot_absatz,
                "predicted_nicht_gedeckt": pred_nicht_gedeckt,
            }

        if not competitor_predictions:
            return None

        preise = [v["predicted_preis"] for v in competitor_predictions.values()]
        predicted_branche_avg_preis = sum(preise) / len(preise) if preise else 3000
        predicted_branche_avg_kz = sum(
            team_preds[u]["pred_kz"] for u in competitor_predictions
        ) / max(1, len(competitor_predictions))
        predicted_branche_nicht_gedeckt = sum(
            v["predicted_nicht_gedeckt"] for v in competitor_predictions.values()
        )

        return {
            "predicted_branche_avg_preis": predicted_branche_avg_preis,
            "predicted_branche_avg_kz": predicted_branche_avg_kz,
            "predicted_branche_nicht_gedeckt": predicted_branche_nicht_gedeckt,
            "competitor_details": competitor_predictions,
        }

    def optimiere_entscheidungen(self, ziel="balanced", target_periode=None, pred_mode="advanced"):
        """Findet die rechnerisch besten Entscheidungen per Optimierung.
        ziel: 'balanced' (gewichteter Mix), 'aktienkurs', 'ebit', 'eigenkapital'
        pred_mode: 'advanced' (MCI) oder 'simple' (Trend)
        """
        tp = target_periode or (self.state["periode"] + 1)
        s_backup = deepcopy(self.state)

        # --- V6.2: Competitor Prediction Integration (MCI / Simple) ---
        predictions = self.predict_competitors(tp, mode=pred_mode)
        if predictions:
            print(f"  [V6 AI] Gegner-Vorhersage ({pred_mode}): ø-Preis={predictions['predicted_branche_avg_preis']:.0f} EUR, ø-KZ={predictions.get('predicted_branche_avg_kz', 60):.1f}")
            s_backup["branche_avg_preis"] = predictions["predicted_branche_avg_preis"]
            s_backup["branche_avg_kz"] = predictions.get("predicted_branche_avg_kz", 60)
            s_backup["branche_nicht_gedeckt"] = predictions["predicted_branche_nicht_gedeckt"]
            current_avg = self.state.get("branche_avg_preis", 3000)
            if current_avg > 0:
                preis_delta_ratio = (predictions["predicted_branche_avg_preis"] - current_avg) / current_avg
                pot_absatz_adjust = 1.0 - preis_delta_ratio * 0.5
                s_backup["pot_absatz_vor"] = max(
                    1000,
                    s_backup.get("pot_absatz_vor", 43000) * pot_absatz_adjust
                )
        # -----------------------------------------------

        markt2_offen = self._is_markt2_open_for_period(tp)
        s_backup["markt2_offen"] = markt2_offen

        print(f"\n  Optimiere Entscheidungen fuer P{tp} (Ziel: {ziel})...")
        if self.history.period_count() < 4:
            print("  Hinweis: <4 Perioden Historie, Optimierung ist nur als grobe Orientierung zu verstehen.")

        def objective(trial):
            d = s_backup.copy()
            # Fallbacks für Variablen, die nicht aktiv optimiert werden
            d["neue_anlagen_a"] = 0
            d["neue_anlagen_b"] = 0
            d["neue_anlagen_c"] = 0
            d["personal_aenderung_fert"] = 0
            d["fe_personal_aenderung"] = 0
            d["markt2_aktiv"] = s_backup.get("markt2_offen", False)

            d["preis_m1"] = trial.suggest_float("preis_m1", 2800, 4500)
            d["werbung_m1"] = trial.suggest_float("werbung_m1", 0.0, 20.0)
            d["vertrieb_ma"] = trial.suggest_int("vertrieb_ma", 50, 180)
            d["fertigungsmenge"] = trial.suggest_int("fertigungsmenge", 20000, 60000)
            d["personal_fe"] = trial.suggest_int("personal_fe", 25, 90)
            d["personal_fe_gen2"] = trial.suggest_int("personal_fe_gen2", 0, 60)
            d["oeko_budget_gen1"] = trial.suggest_float("oeko_budget_gen1", 0.0, 5.0)
            d["oeko_budget_gen2"] = trial.suggest_float("oeko_budget_gen2", 0.0, 3.0)
            d["ci_budget"] = trial.suggest_float("ci_budget", 0.0, 3.0)
            d["rationalisierung"] = trial.suggest_float("rationalisierung", 0.0, 5.0)
            d["wertanalyse"] = trial.suggest_float("wertanalyse", 0.0, 2.0)
            d["grossabnehmer"] = trial.suggest_int("grossabnehmer", 0, 15000)

            if s_backup.get("markt2_offen", False):
                d["preis_m2"] = trial.suggest_float("preis_m2", 20000, 56000)
                d["werbung_m2"] = trial.suggest_float("werbung_m2", 0.0, 10.0)
                d["vertrieb_m2"] = trial.suggest_int("vertrieb_m2", 0, 50)

            # Bestehende Berechnung aufrufen (Rückgabewert negieren für minimize)
            result = self._berechne(d, s_backup)

            # Ziel-Metrik dynamisch auswerten
            if ziel == "ebit":
                score = result.get("ebit", 0.0)
            elif ziel == "aktienkurs":
                score = result.get("aktienkurs", 0.0)
            elif ziel == "eigenkapital":
                score = result.get("eigenkapital", 0.0)
            else:  # "balanced"
                score = (
                    result.get("aktienkurs", 100) * 0.3 +
                    result.get("ebit", 0) * 5.0 * 0.3 +
                    result.get("mva", 0) * 0.2 +
                    result.get("kz_index", 60) * 2.0 * 0.2
                )

            return -float(score)

        print("  Starte KI-gestützte Bayesian Optimization (Optuna) mit 300 Trials...")
        study = optuna.create_study(direction="minimize")
        study.optimize(objective, n_trials=300)

        print(f"  {'Entscheidung':<25} {'Optimal':>12}  {'Aktuell':>12}  {'Aenderung':>12}")
        print("  " + "-"*25 + " " + "-"*12 + " " + "-"*12 + " " + "-"*12)

        best_decisions = s_backup.copy()
        best_decisions["neue_anlagen_a"] = 0
        best_decisions["neue_anlagen_b"] = 0
        best_decisions["personal_aenderung_fert"] = 0
        best_decisions["fe_personal_aenderung"] = 0
        best_decisions["markt2_aktiv"] = s_backup.get("markt2_offen", False)

        for key, val in study.best_params.items():
            best_decisions[key] = val
            akt = s_backup.get(key, 0)
            diff = val - akt

            if isinstance(val, int) or key.endswith("menge") or key.endswith("ma") or "personal" in key:
                print(f"  {key:<25} {int(val):>12}  {int(akt):>12}  {int(diff):>12}")
            else:
                print(f"  {key:<25} {val:>12.2f}  {akt:>12.2f}  {diff:>12.2f}")

        erg = self._berechne_mit_news(best_decisions, state_override=s_backup, target_period=tp)

        print(f"\n  Prognostizierte Ergebnisse:")
        print(f"  {'Kennzahl':30s} {'Wert':>12s}")
        print(f"  {'-'*30} {'-'*12}")
        kpi_list = [
            ("Aktienkurs", erg["aktienkurs"], ""),
            ("Pot. Absatz", erg["pot_absatz"], "Stk"),
            ("Umsatz", erg["umsatz"], "MEUR"),
            ("EBIT", erg["ebit"], "MEUR"),
            ("Per.Ueberschuss", erg["periodenueberschuss"], "MEUR"),
            ("Eigenkapital", erg["eigenkapital"], "MEUR"),
            ("MVA", erg["mva"], "MEUR"),
            ("EVA", erg["eva"], "MEUR"),
            ("KZ-Index", erg["kz_index"], ""),
            ("Tech-Index", erg["tech_index"], ""),
            ("Ueberziehung", erg["ueberziehung"], "MEUR"),
            ("Lager", erg["neues_lager"], "Stk"),
        ]
        for label, val, unit in kpi_list:
            if isinstance(val, float):
                print(f"  {label:30s} {val:>12.2f} {unit}")
            else:
                print(f"  {label:30s} {val:>12.0f} {unit}")
        print(f"{'='*78}")

        return best_decisions, erg

    def optimiere_vergleich(self, target_periode=None):
        """Optimiert fuer verschiedene Ziele und zeigt Vergleich."""
        tp = target_periode or (self.state["periode"] + 1)
        ziele = ["balanced", "aktienkurs", "ebit", "eigenkapital"]
        ergebnisse = {}
        for z in ziele:
            res = self.optimiere_entscheidungen(ziel=z, target_periode=tp)
            if res:
                ergebnisse[z] = res

        if len(ergebnisse) > 1:
            print(f"\n{'='*100}")
            print(f"  VERGLEICH OPTIMIERUNGSZIELE P{tp}")
            print(f"{'='*100}")
            header = f"  {'Kennzahl':22s}"
            for z in ziele:
                if z in ergebnisse:
                    header += f" {z:>14s}"
            print(header)
            print(f"  {'-'*22}" + f" {'-'*14}" * len(ergebnisse))

            for kpi in ["aktienkurs", "ebit", "eigenkapital", "mva", "periodenueberschuss", "kz_index"]:
                row = f"  {kpi:22s}"
                for z in ziele:
                    if z in ergebnisse:
                        val = ergebnisse[z][1].get(kpi, 0)
                        row += f" {val:>14.2f}"
                print(row)
            print(f"{'='*100}")

        return ergebnisse

    # ---- Entscheidungseingabe ----

    def input_decisions(self):
        target_p = self.state["periode"] + 1
        markt2_offen = self._is_markt2_open_for_period(target_p)
        print(f"=== Entscheidungen fuer Periode {target_p} ===\n")
        d = {}
        s = self.state
        d["preis_m1"] = _prompt_float(
            f"  Preis Markt 1 (EUR) [{self.state['preis_vor']}]: ",
            default=self.state["preis_vor"], min_val=2000, max_val=7000
        )
        d["werbung_m1"] = _prompt_float(
            f"  Werbung Markt 1 (MEUR) [{self.state['werbung_vor']}]: ",
            default=self.state["werbung_vor"], min_val=0
        )
        d["vertrieb_ma"] = _prompt_int(
            f"  Vertriebs-MA [{self.state['personal_vertrieb']}]: ",
            default=self.state["personal_vertrieb"], min_val=0
        )
        d["einkauf_menge"] = _prompt_float(f"  Einkauf Einsatzstoffe (Stk) [{s.get('einkauf_menge', 45000)}]: ", default=s.get("einkauf_menge", 45000))
        d["fertigungsmenge"] = _prompt_int(
            "  Fertigungsmenge Gen1 (Stk) [50000]: ",
            default=50000, min_val=0
        )
        d["personal_fe"] = _prompt_float(f"  Personal F&E Gen 1 (Anzahl) [{s.get('personal_fe', 35)}]: ", default=s.get("personal_fe", 35))
        d["personal_fe_gen2"] = _prompt_float(f"  Personal F&E Gen 2 (Anzahl) [{s.get('personal_fe_gen2', 0)}]: ", default=s.get("personal_fe_gen2", 0))
        d["oeko_budget_gen1"] = _prompt_float(f"  Oekologie-Budget Gen 1 (MEUR) [{s.get('oeko_budget_gen1', 0.0)}]: ", default=s.get("oeko_budget_gen1", 0.0))
        d["oeko_budget_gen2"] = _prompt_float(f"  Oekologie-Budget Gen 2 (MEUR) [{s.get('oeko_budget_gen2', 0.0)}]: ", default=s.get("oeko_budget_gen2", 0.0))
        d["ci_budget"] = _prompt_float(f"  Corporate Identity (MEUR) [{s.get('ci_budget', 0.0)}]: ", default=s.get("ci_budget", 0.0))
        d["rationalisierung"] = _prompt_float("  Rationalisierung (MEUR) [0]: ", default=0, min_val=0)
        d["wertanalyse"] = _prompt_float("  Wertanalyse (MEUR) [0]: ", default=0, min_val=0)
        d["neue_anlagen_a"] = _prompt_int("  Neue Typ A Anlagen [0]: ", default=0, min_val=0)
        d["neue_anlagen_b"] = _prompt_int("  Neue Typ B Anlagen [0]: ", default=0, min_val=0)
        d["personal_aenderung_fert"] = _prompt_int("  Personal Fertigung +/- [0]: ", default=0)
        d["grossabnehmer"] = _prompt_float(f"  Grossabnehmer Menge (Stk) [{s.get('grossabnehmer', 0)}]: ", default=s.get("grossabnehmer", 0))
        d["plan_absatz_m1"] = _prompt_float(f"  Plan-Absatz M1 (Stk) [{s.get('plan_absatz_m1', 45000)}]: ", default=s.get("plan_absatz_m1", 45000))
        d["plan_absatz_m2"] = _prompt_float(f"  Plan-Absatz M2 (Stk) [{s.get('plan_absatz_m2', 0)}]: ", default=s.get("plan_absatz_m2", 0))
        mf = input(f"  Marktforschungsbericht kaufen? (j/n) [{'j' if s.get('marktforschung_aktiv') else 'n'}]: ").strip().lower()
        d["marktforschung_aktiv"] = (mf == 'j' or (mf == '' and s.get("marktforschung_aktiv", False)))
        d["markt2_aktiv"] = False
        if markt2_offen:
            d["markt2_aktiv"] = _prompt_yes_no("  Markt 2 aktiv? (j/n)", default=False)
        if d["markt2_aktiv"]:
            d["preis_m2"] = _prompt_float("  Preis Markt 2 (FCU) [28000]: ", default=28000, min_val=10000, max_val=80000)
            d["werbung_m2"] = _prompt_float("  Werbung Markt 2 (MEUR) [0]: ", default=0, min_val=0)
            d["vertrieb_m2"] = _prompt_float(f"  Vertriebs-MA Markt 2 [{s.get('vertrieb_m2', 0)}]: ", default=s.get("vertrieb_m2", 0))
        else:
            d["preis_m2"] = d["werbung_m2"] = 0
        d["kredit"] = _prompt_float("  Neuer Kredit (MEUR) [0]: ", default=0, min_val=0)
        d["dividende"] = _prompt_float("  Dividende (MEUR) [0]: ", default=0, min_val=0)
        self.decisions = d
        print()

    def _einkaufspreis(self, menge):
        for grenze, preis in self.calibration.params["einkauf_staffel"]:
            if menge <= grenze:
                return preis
        return 400

    # ---- Kernberechnung (V5.1 mit allen Handbuch-Mechanismen) ----

    def _berechne(self, decisions, state_override=None):
        d = decisions.copy()
        s = state_override or self.state
        p = self.calibration.params
        next_p = s["periode"] + 1

        # --- A4/B1: Produktivitaet & Personal ---
        personal_fert_neu = s["personal_fert"] + d.get("personal_aenderung_fert", 0)
        fehlzeiten_rate = p["fehlzeiten_pct"]
        einsetzbares_personal = personal_fert_neu * (1 - fehlzeiten_rate)
        prod_idx1 = s.get("produktivitaet_idx1", 1.0)
        kumul_fert = s.get("kumul_fertigung", self.BASIS_KUMUL_FERTIGUNG) + d["fertigungsmenge"]
        prod_idx2 = 1.0 + math.log(max(kumul_fert, self.BASIS_KUMUL_FERTIGUNG) / self.BASIS_KUMUL_FERTIGUNG) * 0.03
        tats_produktivitaet = p["grundproduktivitaet"] * prod_idx1 * prod_idx2
        personal_kapazitaet = einsetzbares_personal * tats_produktivitaet
        fertigungsmenge_gen2 = d.get("fertigungsmenge_gen2", 0)
        fertigungsmenge_gen1 = d.get("fertigungsmenge", 0) - fertigungsmenge_gen2
        benoetigte_kapazitaet = fertigungsmenge_gen1 * 1 + fertigungsmenge_gen2 * 2
        fertigungs_mitarbeiter_bedarf = (fertigungsmenge_gen1 / 50) + (fertigungsmenge_gen2 / 80)

        # --- B3: Rationalisierung ---
        ration_invest = d.get("rationalisierung", 0)
        ration_idx = s.get("rationalisierung_index", 1.0) + ration_invest * p["ration_index_per_meur"]

        # --- Anlagen-Kapazitaet ---
        anlagen_kap_basis = (
            s["anlagen_kapazitaet"]
            + d["neue_anlagen_a"] * self.ANLAGEN_KAP_A
            + d["neue_anlagen_b"] * p.get("anlagen_kap_b", self.ANLAGEN_KAP_B)
            + d.get("neue_anlagen_c", 0) * p.get("anlagen_kap_c", 25000)
        )
        anlagen_kap = anlagen_kap_basis * ration_idx

        # --- A3: Ueberstunden ---
        engpass_anlagen = benoetigte_kapazitaet > anlagen_kap
        engpass_personal = fertigungs_mitarbeiter_bedarf > einsetzbares_personal
        ueberstunden_aktiv = engpass_anlagen or engpass_personal
        max_kap_anlagen = anlagen_kap * (1 + p["ueberstunden_max_pct"]) if engpass_anlagen else anlagen_kap
        max_kap_personal = personal_kapazitaet * (1 + p["ueberstunden_max_pct"]) if engpass_personal else personal_kapazitaet
        tats_benoetigte_kapazitaet = min(benoetigte_kapazitaet, max_kap_anlagen, max_kap_personal)
        kapazitaetsquote = tats_benoetigte_kapazitaet / max(benoetigte_kapazitaet, 1)
        tats_fertigungsmenge = d["fertigungsmenge"] * kapazitaetsquote

        ueberstunden_kosten = 0.0
        ueberstunden_lohnzuschlag = 0.0
        if ueberstunden_aktiv and tats_benoetigte_kapazitaet > min(anlagen_kap, personal_kapazitaet):
            ueberstunden_kosten = p["ueberstunden_sprungfix"]
            if engpass_personal:
                ueberstunden_anteil = (tats_benoetigte_kapazitaet - personal_kapazitaet) / max(personal_kapazitaet, 1)
                ueberstunden_lohnzuschlag = s.get("loehne_basis", 25.14) * ueberstunden_anteil * p["ueberstunden_lohnzuschlag"]

        # --- TECH (Investment + Personal) ---
        fe_change = d.get("fe_personal_aenderung")
        if fe_change is None:
            fe_change = d.get("gen2_personal", 0)
        fe_personal = s.get("personal_fe", 34) + fe_change
        fe_personal_effekt = max(0, fe_personal - p.get("fe_personal_basis", 30)) * p.get("tech_per_fe_ma", 0.1)

        # F&E Investment ergibt sich jetzt aus den Loehnen der F&E-Mitarbeiter
        personal_fe = d.get("personal_fe", s.get("personal_fe", 35.0))
        fe_basis_gehalt = 45000.0 * (p.get("loehne_steigerung", 1.03) ** s.get("periode", 0))
        fe_loehne_meur = (personal_fe * fe_basis_gehalt) / 1_000_000.0

        # Neue Entscheidungsfelder aus TOPSIM V6.2
        personal_fe_gen2 = d.get("personal_fe_gen2", s.get("personal_fe_gen2", 0))
        oeko_budget_gen1 = d.get("oeko_budget_gen1", s.get("oeko_budget_gen1", 0.0))
        oeko_budget_gen2 = d.get("oeko_budget_gen2", s.get("oeko_budget_gen2", 0.0))
        ci_budget = d.get("ci_budget", s.get("ci_budget", 0.0))
        grossabnehmer = d.get("grossabnehmer", s.get("grossabnehmer", 0))
        einkauf_menge = d.get("einkauf_menge", s.get("einkauf_menge", d.get("fertigungsmenge", 45000)))
        marktforschung_aktiv = d.get("marktforschung_aktiv", False)
        vertrieb_m2 = d.get("vertrieb_m2", s.get("vertrieb_m2", 0))

        fe_invest_gen1 = fe_loehne_meur
        d["fe_invest_gen1"] = fe_invest_gen1  # Wird fuer tech_neu und die GuV weiterverwendet

        tech_neu = s["tech_index"] + p["tech_per_meur"] * d["fe_invest_gen1"] + fe_personal_effekt

        # --- A8: Umweltindex ---
        umweltindex = s.get("umweltindex", self.BASIS_UMWELTINDEX)
        neue_kap_fuer_umwelt = (
            d["neue_anlagen_a"] * self.ANLAGEN_KAP_A
            + d.get("neue_anlagen_b", 0) * p.get("anlagen_kap_b", self.ANLAGEN_KAP_B)
            + d.get("neue_anlagen_c", 0) * p.get("anlagen_kap_c", 25000)
        )
        if neue_kap_fuer_umwelt > 0:
            alte_kap = s["anlagen_kapazitaet"]
            umweltindex = (umweltindex * alte_kap + 100.0 * neue_kap_fuer_umwelt) / max(alte_kap + neue_kap_fuer_umwelt, 1)
        oeko_verbesserung = d.get("oeko_budget", 0) * 1.5
        umwelt_bonus = (oeko_budget_gen1 + oeko_budget_gen2) * 0.5
        umweltindex = min(100.0, umweltindex + oeko_verbesserung + umwelt_bonus)
        umwelt_strafe_meur = umwelt_strafe(umweltindex)

        # --- A5: Bekanntheit ---
        bekanntheit_alt = s.get("bekanntheit", 49.25)
        ci_effekt = d.get("ci_budget", 0) * 1.5
        ci_bonus = ci_budget * 0.3
        werbe_effekt = (d["werbung_m1"] - s["werbung_vor"]) * 0.3
        bekanntheit_neu = bekanntheit_alt + ci_effekt + ci_bonus + werbe_effekt
        bekanntheit_neu = max(30, min(90, bekanntheit_neu))

        # --- POT. ABSATZ: Konsistenter Nakanishi-Cooper Hybrid (V6.2) ---
        # Nutzt pot_absatz_vor als robusten Anker und korrigiert über die relative Preisposition
        preis_ratio = d["preis_m1"] / max(s["preis_vor"], 1)
        werb_ratio = d["werbung_m1"] / max(s["werbung_vor"], 0.1)
        tech_ratio = tech_neu / max(s["tech_index"], 1)
        vertrieb_ratio = d["vertrieb_ma"] / max(s.get("personal_vertrieb", 100), 1)
        bekanntheit_ratio = bekanntheit_neu / max(bekanntheit_alt, 1)

        branche_avg = s.get("branche_avg_preis", 3000)
        preis_relativ = d["preis_m1"] / max(branche_avg, 1)
        branche_avg_kz = max(1.0, s.get("branche_avg_kz", 60.0))
        kz_relativ = s["kz_index"] / branche_avg_kz

        pot_m1 = s["pot_absatz_vor"] \
            * (preis_ratio ** p.get("price_elasticity", -1.6)) \
            * (preis_relativ ** p.get("preis_relativ_exp", -0.85)) \
            * (werb_ratio ** p.get("werbung_exponent", 0.1)) \
            * (tech_ratio ** p.get("tech_exponent", 0.3)) \
            * (vertrieb_ratio ** p.get("vertrieb_exponent", 0.15)) \
            * (bekanntheit_ratio ** p.get("bekanntheit_exponent", 0.08)) \
            * (kz_relativ ** p.get("kz_exponent", 0.1))

        spillover = pot_m1 * p["spillover_sweetspot"] if 3200 <= d["preis_m1"] <= 3700 else 0
        pot_m1 += spillover

        # --- MARKET CLEARING: Lieferunfähigkeit & Umverteilung (Handbuch-konform) ---
        fremde_ungedeckt = 0.0
        eigene_ungedeckt = s.get("nicht_gedeckt_vor", 0)

        if "competitor_predictions" in d:
            for u, comp in d["competitor_predictions"].items():
                if u != f"U{self.u_nr}":
                    fremde_ungedeckt += comp.get("predicted_nicht_gedeckt", 0.0)
        elif "alle_unternehmen" in d:  # Backtest-Modus
            own_u = f"U{self.u_nr}"
            for u, comp in d["alle_unternehmen"].items():
                if u != own_u:
                    fremde_ungedeckt += max(0.0, float(comp.get("nicht_gedeckt", 0.0)))
        else:
            # Dynamischer Intra-Perioden Fallback (Standard-Simulation)
            target_p = s.get("periode", 0) + 1
            preds = self.predict_competitors(target_p, mode="advanced")
            if preds and "competitor_details" in preds:
                for u, comp in preds["competitor_details"].items():
                    if u != f"U{self.u_nr}":
                        fremde_ungedeckt += comp.get("predicted_nicht_gedeckt", 0.0)

        branche_pot_summe = max(1.0, float(s.get("branche_pot_absatz_summe", 250000)))
        mein_anteil_an_umverteilung = pot_m1 / branche_pot_summe
        spillover_nachfrage = fremde_ungedeckt * 0.8 * mein_anteil_an_umverteilung
        spillover_liefer = spillover_nachfrage  # Alias fuer Rueckgabe-Dict
        nachfrage_gesamt = pot_m1 + spillover_nachfrage

        # --- A6: Kundenzufriedenheit (vollstaendig) ---
        kz_delta_preis = (d["preis_m1"] - s["preis_vor"]) / 100 * p["kz_price_sensitivity"]
        lager_alt_ratio = min(s["lager_fertig"] / max(pot_m1, 1) * 100, 30)
        kz_delta_lager = lager_alt_ratio * p["kz_lager_penalty"]
        kz_delta_umwelt = max(0, umweltindex - self.BASIS_UMWELTINDEX) * p["kz_umwelt_bonus"]
        kz_delta_liefer = p["kz_lieferunfaehig_penalty"] if eigene_ungedeckt > 0 else 0
        kz_neu = s["kz_index"] + kz_delta_preis + kz_delta_lager + kz_delta_umwelt + kz_delta_liefer
        kz_neu = max(20, min(95, kz_neu))

        # --- ABSATZ & LAGER ---
        target_period = s["periode"] + 1
        markt2_offen = self._is_markt2_open_for_period(target_period)
        markt2_aktiv = bool(d.get("markt2_aktiv", False)) and markt2_offen

        supply = tats_fertigungsmenge + s["lager_fertig"]
        gross = min(grossabnehmer, self.MAX_GROSSABNEHMER)
        pot_m1_mit_spillover = nachfrage_gesamt

        pot_m2 = 0.0
        if markt2_aktiv:
            m2_pot_vor = s.get("pot_absatz_m2_vor", 0)
            if m2_pot_vor <= 0:
                m2_pot_vor = p.get("markt2_start_potenzial", 12000)
            preis_m2_vor = max(1, s.get("preis_m2_vor", d.get("preis_m2", 4200)))
            werbung_m2_vor = max(0.1, s.get("werbung_m2_vor", max(d.get("werbung_m2", 0), 0.1)))
            preis_ratio_m2 = max(1e-6, d.get("preis_m2", 4200) / preis_m2_vor)
            werb_ratio_m2 = max(1e-6, max(d.get("werbung_m2", 0), 0.1) / werbung_m2_vor)
            vertrieb_ratio_m2 = max(1e-6, max(vertrieb_m2, 1) / max(s.get("vertrieb_m2", max(vertrieb_m2, 1)), 1))
            pot_m2 = (
                m2_pot_vor
                * (preis_ratio_m2 ** p["price_elasticity"])
                * (werb_ratio_m2 ** p["werbung_exponent"])
                * (tech_ratio ** (p["tech_exponent"] * 0.8))
                * (vertrieb_ratio_m2 ** p.get("vertrieb_exponent", 0.15))
            )
            pot_m2 = max(0.0, pot_m2)

        freie_kap = max(0.0, supply - gross)
        demand_markets = max(0.0, pot_m1_mit_spillover) + max(0.0, pot_m2)
        if demand_markets <= 0 or freie_kap <= 0:
            tats_m1 = 0.0
            tats_m2 = 0.0
        else:
            tats_m1 = min(pot_m1_mit_spillover, freie_kap * (pot_m1_mit_spillover / demand_markets))
            tats_m2 = min(pot_m2, max(0.0, freie_kap - tats_m1))
            rest = max(0.0, freie_kap - tats_m1 - tats_m2)
            if rest > 0 and pot_m1_mit_spillover > tats_m1:
                extra = min(rest, pot_m1_mit_spillover - tats_m1)
                tats_m1 += extra
                rest -= extra
            if rest > 0 and pot_m2 > tats_m2:
                tats_m2 += min(rest, pot_m2 - tats_m2)

        tats_gesamt = gross + tats_m1 + tats_m2
        neues_lager = max(0, supply - tats_gesamt)
        nicht_gedeckt_m1 = max(0, pot_m1_mit_spillover - tats_m1)
        nicht_gedeckt_m2 = max(0, pot_m2 - tats_m2)
        nicht_gedeckt = nicht_gedeckt_m1 + nicht_gedeckt_m2

        # --- B4: Wertanalyse (Materialeinsparung) ---
        wertanalyse_idx = p.get("wertanalyse_index", 1.0) + d.get("wertanalyse", 0) * 0.02
        material_bedarf = tats_fertigungsmenge / wertanalyse_idx

        # --- KOSTEN ---
        einkauf_preis = self._einkaufspreis(einkauf_menge)
        material_regulaer = einkauf_menge * einkauf_preis / 1e6
        fehlmenge_material = max(0, material_bedarf - einkauf_menge)
        sonderkosten_material = fehlmenge_material * p.get("luftfracht_preis", 780) / 1e6
        material_var = material_regulaer + sonderkosten_material

        ueberschuss_material = max(0, einkauf_menge - material_bedarf)
        lagerkosten_einsatz = ueberschuss_material * p.get("lagerkosten_einsatz_stueck", 50) / 1e6
        betriebs_var = tats_fertigungsmenge * p["betriebsstoff_stueck"] / 1e6
        transport_m1 = tats_m1 * p["transport_stueck"] / 1e6
        transport_m2 = tats_m2 * p.get("transport_m2_stueck", 75) / 1e6
        transport = transport_m1 + transport_m2

        # A2: Nacharbeit
        nacharbeit_pct = p["nacharbeit_basis_pct"] + max(0, tech_neu - 100) * p["nacharbeit_tech_factor"]
        nacharbeit = (material_var + betriebs_var) * nacharbeit_pct / 100

        sozial = p.get("sozialkosten_faktor", 1.48)
        lohn_fert_teur = p.get("lohn_fertigung_teur", 31) * sozial
        loehne = personal_fert_neu * lohn_fert_teur / 1e3
        loehne += ueberstunden_lohnzuschlag

        personal_vertrieb = d.get("vertrieb_ma", s.get("personal_vertrieb", 100))
        personal_fe = fe_personal

        # Fix 2: Verwaltung + Einkauf automatisch umsatzabhaengig
        umsatz_schaetzung = s.get("pot_absatz_vor", 43000) * d["preis_m1"] / 1e6
        personal_admin = round(_interpolate_kurve(VERWALTUNG_KURVE, umsatz_schaetzung))
        personal_einkauf = round(_interpolate_kurve(EINKAUF_KURVE, umsatz_schaetzung))

        gehaelter = (personal_vertrieb * p.get("gehalt_vertrieb", 41) * sozial
                     + personal_fe * p.get("gehalt_fe", 45) * sozial
                     + personal_admin * p.get("gehalt_admin", 29) * sozial
                     + personal_einkauf * p.get("gehalt_einkauf", 31) * sozial) / 1e3

        # Fix 1: Fluktuation (~6%) -> natuerlicher Abgang, muss ersetzt werden
        flukt_rate = p.get("fluktuation_rate", 0.058)
        flukt_fert = round(s.get("personal_fert", 852) * flukt_rate)
        flukt_vert = round(s.get("personal_vertrieb", 100) * flukt_rate)
        flukt_fe = round(s.get("personal_fe", 34) * flukt_rate)
        flukt_admin = round(personal_admin * flukt_rate)
        flukt_einkauf = round(personal_einkauf * flukt_rate)
        flukt_gesamt = flukt_fert + flukt_vert + flukt_fe + flukt_admin + flukt_einkauf

        personal_aenderung_fert = d.get("personal_aenderung_fert", 0)
        personal_aenderung_vert = d.get("vertrieb_ma", personal_vertrieb) - s.get("personal_vertrieb", personal_vertrieb)
        personal_aenderung_fe = fe_change

        einstell_k = p.get("einstellungskosten_teur", 5) / 1e3
        entlass_k = p.get("entlassungskosten_teur", 10) / 1e3

        einst_fert_plan = max(0, personal_aenderung_fert)
        entl_fert_plan = max(0, -personal_aenderung_fert)
        einst_vert_plan = max(0, personal_aenderung_vert)
        entl_vert_plan = max(0, -personal_aenderung_vert)
        einst_fe_plan = max(0, personal_aenderung_fe)
        entl_fe_plan = max(0, -personal_aenderung_fe)

        einstellungen_gesamt = einst_fert_plan + einst_vert_plan + einst_fe_plan + flukt_gesamt
        entlassungen_fert = entl_fert_plan
        entlassungen_vert = entl_vert_plan
        entlassungen_fe = entl_fe_plan

        personalwechsel_kosten = einstellungen_gesamt * einstell_k
        personalwechsel_kosten += (entlassungen_fert + entlassungen_vert + entlassungen_fe) * entlass_k

        # Fix 3: Sozialplan bei Massenentlassungen (>10%)
        sozialplan = 0.0
        if entlassungen_fert > 0:
            sp_k = sozialplan_kosten_pro_person(entlassungen_fert, s.get("personal_fert", 852))
            sozialplan += entlassungen_fert * sp_k / 1e3
        if entlassungen_vert > 0:
            sp_k = sozialplan_kosten_pro_person(entlassungen_vert, s.get("personal_vertrieb", 100))
            sozialplan += entlassungen_vert * sp_k / 1e3
        personalwechsel_kosten += sozialplan

        personalaufwand = loehne + gehaelter + personalwechsel_kosten

        werbung_ges = d["werbung_m1"] + d.get("werbung_m2", 0)
        fe_ges = d["fe_invest_gen1"] + d.get("oeko_budget", 0) + d.get("ci_budget", 0) + d.get("wertanalyse", 0)
        fix = p["fixkosten_basis"] + d.get("rationalisierung", 0)
        if marktforschung_aktiv:
            fix += 0.1
        fix += d["neue_anlagen_a"] * 1.25 + d.get("neue_anlagen_b", 0) * p.get("anlagen_fix_b", 1.0) + d.get("neue_anlagen_c", 0) * p.get("anlagen_fix_c", 6.0)
        fix += ueberstunden_kosten
        anlagen_anzahl = s.get("anlagen_anzahl", 4)
        instandhaltung_wartung = anlagen_anzahl * 1.0
        instandhaltung_rat = d.get("instandhaltung_rat", 0.0)
        instandhaltung_gesamt = instandhaltung_wartung + instandhaltung_rat
        fix += p["gebaeude_abschreibung"] + instandhaltung_gesamt

        # A7: Lagerkosten (Fertigerzeugnisse + Einsatzstoffe)
        avg_lager = (s["lager_fertig"] + neues_lager) / 2
        lagerkosten = (avg_lager * p["lagerkosten_fertig_stueck"] / 1e6) + lagerkosten_einsatz

        # A8: Umweltstrafe
        sonstiger_aufwand = umwelt_strafe_meur + nacharbeit + lagerkosten

        # --- UMSATZ ---
        umsatz_m1 = tats_m1 * d["preis_m1"] / 1e6
        umsatz_m2 = tats_m2 * d.get("preis_m2", 0) * p.get("wechselkurs", 0.125) / 1e6
        umsatz_gross = gross * p["grossabnehmer_preis"] / 1e6
        umsatz = umsatz_m1 + umsatz_m2 + umsatz_gross

        # --- ERGEBNIS ---
        kosten = material_var + betriebs_var + transport + werbung_ges + fe_ges + personalaufwand + fix + s["abschreibungen"] + sonstiger_aufwand
        herstellkosten_gesamt = material_var + betriebs_var + loehne + nacharbeit + ueberstunden_kosten + ueberstunden_lohnzuschlag
        herstellkosten_stueck = herstellkosten_gesamt / max(tats_fertigungsmenge, 1) * 1e6
        neuer_lager_wert = neues_lager * herstellkosten_stueck / 1e6
        alter_lager_wert = s.get("lager_wert", 0)
        if alter_lager_wert <= 0:
            alter_lager_wert = s["lager_fertig"] * herstellkosten_stueck / 1e6
        bestandsaenderung = neuer_lager_wert - alter_lager_wert
        ebit = umsatz + bestandsaenderung - kosten

        # A9: Rating-basierter Zinssatz
        rating_str = s.get("rating", "BB")
        zinssatz = rating_zinssatz(rating_str, p["basiszins"]) / 100
        zinsaufwand = max(0, s["ueberziehung"]) * zinssatz

        gewinn_vor_steuern = ebit - zinsaufwand

        # Fix 5: Verlustvortraege
        verlustvortrag_alt = s.get("verlustvortrag", 0)
        steuerbasis = gewinn_vor_steuern - verlustvortrag_alt
        if steuerbasis > 0:
            steuern = steuerbasis * p["steuersatz"]
            verlustvortrag_neu = 0
        else:
            steuern = 0
            verlustvortrag_neu = abs(steuerbasis)
        netto = gewinn_vor_steuern - steuern
        nopat = ebit * (1 - p["steuersatz"])

        # --- BILANZ ---
        neues_ek = s["eigenkapital"] + netto - d.get("dividende", 0)
        investitionen = (
            d.get("neue_anlagen_a", 0) * p.get("anlagen_preis_a", 21.0) +
            d.get("neue_anlagen_b", 0) * p.get("anlagen_preis_b", 32.0) +
            d.get("neue_anlagen_c", 0) * p.get("anlagen_preis_c", 0.0)
        )
        neues_av = s["anlagevermoegen"] - s["abschreibungen"] + investitionen
        neuer_mva = s["mva"] + nopat * p["mva_delta_faktor"]

        # A1: Aktienkurs mit 11 Faktoren
        umsatzrendite = (netto / max(umsatz, 0.1)) * 100
        fk_quote = max(0, s["ueberziehung"]) / max(neues_ek + s["ueberziehung"], 0.1) * 100
        kumul_div = s.get("kumul_dividende", 0) + d.get("dividende", 0)

        aktien_score = (
            p["aktienkurs_w_ek"] * neues_ek
            + p["aktienkurs_w_netto"] * netto * 5
            + p["aktienkurs_w_umsatzrendite"] * umsatzrendite * 3
            + p["aktienkurs_w_bekanntheit"] * bekanntheit_neu * 0.5
            + p["aktienkurs_w_kz"] * kz_neu * 0.5
            + p["aktienkurs_w_umsatz"] * umsatz * 0.3
            + p["aktienkurs_w_dividende"] * kumul_div * 2
            + p["aktienkurs_w_fkquote"] * fk_quote * 0.3
            + p["aktienkurs_w_techqual"] * tech_neu * 0.3
            + p["aktienkurs_w_mva"] * neuer_mva * 0.5
        )
        aktienkurs = max(10, (neues_ek + neuer_mva) * 2 * 0.5 + aktien_score * 0.5)

        # --- EVA ---
        nce_est = neues_av + 35
        wacc_pct = s.get("wacc", 9.44) / 100
        eva = nopat - wacc_pct * nce_est

        neue_kap_raw = (
            s["anlagen_kapazitaet"]
            + d["neue_anlagen_a"] * self.ANLAGEN_KAP_A
            + d.get("neue_anlagen_b", 0) * p.get("anlagen_kap_b", self.ANLAGEN_KAP_B)
            + d.get("neue_anlagen_c", 0) * p.get("anlagen_kap_c", 25000)
        )

        # B5: Liquiditaetsrechnung (GA 100% sofort, Einzelhandel 80/20)
        einzahlungen = (umsatz_m1 + umsatz_m2) * 0.8 + umsatz_gross + s.get("forderungen_vor", 0)
        neue_forderungen = (umsatz_m1 + umsatz_m2) * 0.2
        neue_ueberziehung = max(0, s["ueberziehung"] - einzahlungen + kosten + zinsaufwand + steuern - d.get("kredit", 0))

        return {
            "periode": next_p, "tech_index": tech_neu,
            "pot_absatz": pot_m1, "pot_absatz_mit_spillover": pot_m1_mit_spillover,
            "pot_absatz_m2": pot_m2,
            "spillover": spillover, "spillover_liefer": spillover_liefer,
            "preis_relativ": preis_relativ,
            "tats_absatz": tats_gesamt, "tats_m1": tats_m1, "tats_m2": tats_m2,
            "grossabnehmer": gross, "neues_lager": neues_lager,
            "nicht_gedeckt": nicht_gedeckt,
            "nicht_gedeckt_m1": nicht_gedeckt_m1,
            "nicht_gedeckt_m2": nicht_gedeckt_m2,
            "kz_index": kz_neu, "bekanntheit": bekanntheit_neu,
            "umweltindex": umweltindex,
            "umsatz": umsatz, "umsatz_m1": umsatz_m1, "umsatz_m2": umsatz_m2, "umsatz_gross": umsatz_gross,
            "material_var": material_var, "betriebs_var": betriebs_var,
            "transport": transport, "transport_m1": transport_m1, "transport_m2": transport_m2,
            "nacharbeit": nacharbeit,
            "loehne": loehne, "gehaelter": gehaelter,
            "personalwechsel_kosten": personalwechsel_kosten,
            "fluktuation": flukt_gesamt, "sozialplan": sozialplan,
            "personal_admin": personal_admin, "personal_einkauf": personal_einkauf,
            "personal_fe": personal_fe,
            "personalaufwand": personalaufwand,
            "werbung_ges": werbung_ges, "fe_ges": fe_ges,
            "fix": fix, "lagerkosten": lagerkosten,
            "umwelt_strafe": umwelt_strafe_meur,
            "sonstiger_aufwand": sonstiger_aufwand,
            "abschreibungen": s["abschreibungen"],
            "herstellkosten_stueck": herstellkosten_stueck,
            "neuer_lager_wert": neuer_lager_wert,
            "bestandsaenderung": bestandsaenderung,
            "ueberstunden_aktiv": ueberstunden_aktiv,
            "ueberstunden_kosten": ueberstunden_kosten,
            "ebit": ebit, "zinsaufwand": zinsaufwand,
            "zinssatz_pct": zinssatz * 100, "rating": rating_str,
            "gewinn_vor_steuern": gewinn_vor_steuern,
            "steuern": steuern, "verlustvortrag": verlustvortrag_neu,
            "periodenueberschuss": netto, "nopat": nopat,
            "eigenkapital": neues_ek, "mva": neuer_mva, "eva": eva,
            "aktienkurs": aktienkurs,
            "umsatzrendite": umsatzrendite, "fk_quote": fk_quote,
            "anlagevermoegen": neues_av,
            "anlagen_kapazitaet": neue_kap_raw,
            "anlagen_kap_effektiv": anlagen_kap,
            "personal_kap": personal_kapazitaet,
            "tats_fertigungsmenge": tats_fertigungsmenge,
            "ueberziehung": neue_ueberziehung,
            "forderungen": neue_forderungen,
            "kumul_fertigung": kumul_fert,
            "rationalisierung_index": ration_idx,
            "preis": d["preis_m1"], "werbung": d["werbung_m1"],
            "markt2_offen": markt2_offen, "markt2_aktiv": markt2_aktiv,
            "preis_m2": d.get("preis_m2", 0), "werbung_m2": d.get("werbung_m2", 0),
            "kumul_dividende": kumul_div,
            "decisions": deepcopy(d),
        }

    def _ergebnis_als_state(self, erg, prev_state=None):
        base = prev_state or self.state
        return {
            "periode": erg["periode"], "tech_index": erg["tech_index"],
            "lager_fertig": erg["neues_lager"], "lager_wert": erg.get("neuer_lager_wert", 0),
            "anlagevermoegen": erg["anlagevermoegen"],
            "ueberziehung": erg["ueberziehung"],
            "eigenkapital": erg["eigenkapital"], "mva": erg["mva"],
            "kz_index": erg["kz_index"],
            "bekanntheit": erg["bekanntheit"],
            "personal_fert": base.get("personal_fert", 850) + erg["decisions"].get("personal_aenderung_fert", 0),
            "personal_vertrieb": erg["decisions"].get("vertrieb_ma", 100),
            "personal_fe": erg.get("personal_fe", base.get("personal_fe", 50)),
            "personal_einkauf": erg.get("personal_einkauf", 20),
            "personal_admin": erg.get("personal_admin", 200),
            "abschreibungen": erg["abschreibungen"],
            "preis_vor": erg["preis"], "werbung_vor": erg["werbung"],
            "pot_absatz_vor": erg["pot_absatz"],
            "pot_absatz_m2_vor": erg.get("pot_absatz_m2", 0),
            "preis_m2_vor": erg.get("preis_m2", base.get("preis_m2_vor", 4200)),
            "werbung_m2_vor": erg.get("werbung_m2", base.get("werbung_m2_vor", 0)),
            "anlagen_kapazitaet": erg["anlagen_kapazitaet"],
            "nopat_vor": erg["nopat"], "wacc": base.get("wacc", 9.44),
            "rating": erg["rating"],
            "umweltindex": erg["umweltindex"],
            "umsatzrendite_vor": erg["umsatzrendite"],
            "fremdkapitalquote_vor": erg["fk_quote"],
            "motivation": base.get("motivation", 50),
            "kumul_fertigung": erg["kumul_fertigung"],
            "produktivitaet_idx1": base.get("produktivitaet_idx1", 1.0),
            "produktivitaet_idx2": 1.0 + math.log(
                max(erg["kumul_fertigung"], self.BASIS_KUMUL_FERTIGUNG) / self.BASIS_KUMUL_FERTIGUNG
            ) * 0.03,
            "nicht_gedeckt_vor": erg["nicht_gedeckt"],
            "kumul_dividende": erg["kumul_dividende"],
            "rationalisierung_index": erg["rationalisierung_index"],
            "forderungen_vor": erg["forderungen"],
            "branche_avg_preis": base.get("branche_avg_preis", 3000),
            "branche_nicht_gedeckt": 0,
            "branche_pot_absatz": base.get("branche_pot_absatz", 258000),
            "loehne_basis": erg["loehne"],
            "personal_summe": (base.get("personal_fert", 850)
                               + erg["decisions"].get("personal_aenderung_fert", 0)
                               + erg["decisions"].get("vertrieb_ma", 100)
                               + erg.get("personal_fe", base.get("personal_fe", 50))
                               + erg.get("personal_einkauf", 20)
                               + erg.get("personal_admin", 200)),
            "verlustvortrag": erg.get("verlustvortrag", 0),
            "sonstiger_aufwand_real": 0,
            "bestandsveraenderung_real": 0,
            "markt2_offen": erg.get("markt2_offen", base.get("markt2_offen", False)),
        }

    def _print_ergebnis(self, erg, compact=False):
        p = self.calibration.params
        if compact:
            print(f"  P{erg['periode']:>2d} | Aktie {erg['aktienkurs']:>6.0f} | "
                  f"Umsatz {erg['umsatz']:>7.2f} | EBIT {erg['ebit']:>7.2f} | "
                  f"EK {erg['eigenkapital']:>7.2f} | MVA {erg['mva']:>7.2f} | "
                  f"Pot.Abs {erg['pot_absatz']:>7.0f} | Lager {erg['neues_lager']:>6.0f} | "
                  f"KZ {erg['kz_index']:>5.1f}")
            return

        ue = " [UE!]" if erg["ueberstunden_aktiv"] else ""
        print(f"{'═'*80}\n  PERIODE {erg['periode']} PROGNOSE – Eagle Eye V6.1 (Auto-Fit + Optimizer)\n{'═'*80}\n")
        print(f"  Tech-Index Gen1       : {erg['tech_index']:.1f}")
        print(f"  Pot. Absatz M1        : {erg['pot_absatz']:,.0f} Stk (Preis-Spill: {erg['spillover']:,.0f} | Liefer-Spill: {erg['spillover_liefer']:,.0f})")
        if erg.get("markt2_offen", False):
            print(f"  Pot. Absatz M2        : {erg.get('pot_absatz_m2', 0):,.0f} Stk")
        print(f"  Preis relativ/Branche : {erg['preis_relativ']:.3f} (1.0 = Branchendurchschnitt)")
        print(
            f"  Tats. Absatz gesamt   : {erg['tats_absatz']:,.0f} Stk "
            f"(M1: {erg['tats_m1']:,.0f} + M2: {erg.get('tats_m2', 0):,.0f} + GA: {erg['grossabnehmer']:,.0f})"
        )
        print(f"  Neues Lager           : {erg['neues_lager']:,.0f} Stk")
        print(f"  KZ-Index              : {erg['kz_index']:.1f}")
        print(f"  Bekanntheit           : {erg['bekanntheit']:.1f}")
        print(f"  Umweltindex           : {erg['umweltindex']:.1f}")
        print(f"  Fertigung             : {erg['tats_fertigungsmenge']:,.0f} Stk{ue}")
        print(f"  Kapazitaet Anlagen    : {erg['anlagen_kap_effektiv']:,.0f} | Personal: {erg['personal_kap']:,.0f}")
        print()
        print(f"  Umsatz                : {erg['umsatz']:.2f} MEUR")
        if erg.get("markt2_offen", False):
            print(f"    - Umsatz M2         : {erg.get('umsatz_m2', 0):.2f} MEUR")
        print(f"  Material variabel     : {erg['material_var']:.2f} MEUR")
        print(f"  Betriebsstoffe        : {erg['betriebs_var']:.2f} MEUR")
        print(f"  Nacharbeit            : {erg['nacharbeit']:.2f} MEUR")
        print(f"  Transport             : {erg['transport']:.2f} MEUR")
        print(f"  Löhne (Fertigung)     : {erg['loehne']:.2f} MEUR")
        print(f"  Gehälter (V/F&E/Verw) : {erg['gehaelter']:.2f} MEUR")
        pw = erg.get('personalwechsel_kosten', 0)
        if pw > 0:
            parts = []
            if erg.get('fluktuation', 0) > 0:
                parts.append(f"Flukt.={erg['fluktuation']}MA")
            if erg.get('sozialplan', 0) > 0:
                parts.append(f"Sozialpl.={erg['sozialplan']:.2f}")
            detail = f" ({', '.join(parts)})" if parts else ""
            print(f"  Einstell/Entlass.kost.: {pw:.2f} MEUR{detail}")
        print(f"  Personal (Verw/Eink)  : {erg.get('personal_admin', 0)}/{erg.get('personal_einkauf', 0)} MA (auto)")
        print(f"  Personalaufwand ges.  : {erg['personalaufwand']:.2f} MEUR")
        print(f"  Werbung               : {erg['werbung_ges']:.2f} MEUR")
        print(f"  F&E + Oeko + CI + WA  : {erg['fe_ges']:.2f} MEUR")
        print(f"  Fixkosten (inkl.UE)   : {erg['fix']:.2f} MEUR{ue}")
        print(f"  Lagerkosten           : {erg['lagerkosten']:.2f} MEUR")
        print(f"  Umwelt-Strafe         : {erg['umwelt_strafe']:.2f} MEUR")
        print(f"  Abschreibungen        : {erg['abschreibungen']:.2f} MEUR")
        print(f"  Bestandsveraenderung  : {erg['bestandsaenderung']:+.2f} MEUR")
        print()
        print(f"  EBIT                  : {erg['ebit']:.2f} MEUR\n")
        print(f"  Zinsen ({erg['zinssatz_pct']:.1f}% | {erg['rating']})  : {erg['zinsaufwand']:.2f} MEUR")
        print(f"  Gewinn vor Steuern    : {erg['gewinn_vor_steuern']:.2f} MEUR")
        vv = erg.get('verlustvortrag', 0)
        vv_txt = f" (VV: {vv:.2f})" if vv > 0 else ""
        print(f"  Steuern ({p['steuersatz']*100:.0f}%)         : {erg['steuern']:.2f} MEUR{vv_txt}")
        print(f"  Periodenueberschuss   : {erg['periodenueberschuss']:.2f} MEUR\n")
        print(f"  NOPAT                 : {erg['nopat']:.2f} MEUR")
        print()
        print(f"  Eigenkapital (neu)    : {erg['eigenkapital']:.2f} MEUR")
        print(f"  MVA (neu)             : {erg['mva']:.2f} MEUR")
        print(f"  EVA (Schaetzung)      : {erg['eva']:.2f} MEUR")
        print(f"  Umsatzrendite         : {erg['umsatzrendite']:.1f}%")
        print(f"  FK-Quote              : {erg['fk_quote']:.1f}%")
        print(f"  ** Aktienkurs         : {erg['aktienkurs']:.0f} EUR **\n")
        print(f"{'='*78}")

    def _print_abweichung(self, erg):
        real = self.history.get_period(erg["periode"])
        if not real:
            return
        print(f"\n  --- Abweichungscheck vs. echtem Report P{erg['periode']} ---")
        checks = [
            ("Aktienkurs", erg["aktienkurs"], real.get("aktienkurs")),
            ("Pot.Absatz", erg["pot_absatz"], real.get("pot_absatz")),
            ("Umsatz", erg["umsatz"], real.get("umsatz_gesamt")),
            ("EBIT", erg["ebit"], real.get("ebit")),
            ("Eigenkapital", erg["eigenkapital"], real.get("eigenkapital")),
            ("MVA", erg["mva"], real.get("mva")),
            ("KZ-Index", erg["kz_index"], real.get("kz_index")),
        ]
        for name, sim_val, real_val in checks:
            if real_val and real_val != 0:
                abw = (sim_val - real_val) / abs(real_val) * 100
                marker = "WARNUNG" if abs(abw) >= 5 else "OK"
                print(f"    {name:20s}: Sim={sim_val:>10.2f} | Real={real_val:>10.2f} | Abw={abw:>+6.1f}% {marker}")
        print()

    def simuliere_periode(self):
        target_p = self.state["periode"] + 1
        erg = self._berechne_mit_news(self.decisions, None, target_p)
        self._print_ergebnis(erg)
        self._print_abweichung(erg)
        self._print_konfidenzband(erg)
        self._scenario_nr += 1
        print(f"  (Szenario #{self._scenario_nr} – State bleibt auf P{self.state['periode']})\n")

    def _print_konfidenzband(self, erg):
        abw = self.geschaetzte_abweichung()
        if not abw:
            return
        n = len(self.history.all_periods()) - 1
        print(f"  --- Geschaetzte Modellgenauigkeit (basierend auf {n} Backtests) ---")
        for name, abw_key, val, unit in [
            ("Aktienkurs", "Aktienkurs", erg["aktienkurs"], "EUR"),
            ("Pot.Absatz", "Pot.Absatz", erg["pot_absatz"], "Stk"),
            ("Umsatz", "Umsatz", erg["umsatz"], "MEUR"),
            ("EBIT", "EBIT", erg["ebit"], "MEUR"),
            ("Eigenkapital", "Eigenkapital", erg["eigenkapital"], "MEUR"),
            ("MVA", "MVA", erg["mva"], "MEUR"),
        ]:
            pct = abw.get(abw_key, 0)
            delta = abs(val * pct / 100)
            print(f"    {name:20s}: {val:>10.1f} {unit:4s}  (+/-{pct:.1f}% -> {val-delta:>10.1f} - {val+delta:>10.1f})")
        gesamt = sum(abw.values()) / len(abw) if abw else 0
        print(f"    {'Modell-Konfidenz':20s}: {max(0, 100 - gesamt * 5):.0f}%")
        print()

    # ---- Multi-Perioden ----

    def simuliere_multi(self, perioden_decisions):
        ergebnisse = []
        temp_state = deepcopy(self.state)
        for dec in perioden_decisions:
            target_p = temp_state["periode"] + 1
            erg = self._berechne_mit_news(dec, temp_state, target_p)
            ergebnisse.append(erg)
            temp_state = self._ergebnis_als_state(erg, prev_state=temp_state)
        return ergebnisse

    def export_ergebnisse(self, ergebnisse, filepath="szenario_export.csv"):
        fields = [
            ("Periode", "periode"), ("Preis M1", "preis"), ("Werbung M1", "werbung"),
            ("Fertigungsmenge", lambda e: e["decisions"]["fertigungsmenge"]),
            ("Grossabnehmer", "grossabnehmer"), ("Tech-Index", "tech_index"),
            ("Pot.Absatz", "pot_absatz"), ("Tats.Absatz", "tats_absatz"),
            ("Lager", "neues_lager"), ("KZ-Index", "kz_index"),
            ("Bekanntheit", "bekanntheit"), ("Umweltindex", "umweltindex"),
            ("Umsatz MEUR", "umsatz"), ("EBIT MEUR", "ebit"), ("NOPAT MEUR", "nopat"),
            ("Periodenueberschuss MEUR", "periodenueberschuss"),
            ("Eigenkapital MEUR", "eigenkapital"), ("MVA MEUR", "mva"),
            ("EVA MEUR", "eva"), ("Aktienkurs EUR", "aktienkurs"),
        ]
        lines = [";".join(name for name, _ in fields)]
        for erg in ergebnisse:
            vals = []
            for _, key in fields:
                if callable(key):
                    vals.append(f"{key(erg):.2f}")
                else:
                    v = erg[key]
                    vals.append(f"{v:.2f}" if isinstance(v, float) else str(v))
            lines.append(";".join(vals))
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        print(f"\n  Exportiert: {filepath} ({len(ergebnisse)} Perioden)")

    # ---- Backtesting & Modellgenauigkeit ----

    def _news_params_for_period(self, target_period):
        """Gibt temporaere Param-Overrides fuer eine Zielperiode zurueck."""
        news = self.get_news(target_period)
        if not news:
            return {}
        overrides = {}
        sozial = self.calibration.params.get("sozialkosten_faktor", 1.48)
        if "lohn_fertigung" in news:
            overrides["lohn_fertigung_teur"] = news["lohn_fertigung"]
        if "lohn_vertrieb" in news:
            overrides["gehalt_vertrieb"] = news["lohn_vertrieb"]
        if "lohn_fe" in news:
            overrides["gehalt_fe"] = news["lohn_fe"]
        if "lohn_verwaltung" in news:
            overrides["gehalt_admin"] = news["lohn_verwaltung"]
        if "lohn_einkauf" in news:
            overrides["gehalt_einkauf"] = news["lohn_einkauf"]
        if "betriebsstoff" in news:
            overrides["betriebsstoff_stueck"] = news["betriebsstoff"]
        if "basiszins" in news:
            overrides["basiszins"] = news["basiszins"]
        if "transport_m1" in news:
            overrides["transport_stueck"] = news["transport_m1"]
        if "transport_m2" in news:
            overrides["transport_m2_stueck"] = news["transport_m2"]
        if "einstellungskosten" in news:
            overrides["einstellungskosten_teur"] = news["einstellungskosten"]
        if "entlassungskosten" in news:
            overrides["entlassungskosten_teur"] = news["entlassungskosten"]
        if "markt2_offen" in news:
            overrides["markt2_offen"] = bool(news["markt2_offen"])
        if "anlagen_kap_b" in news: overrides["anlagen_kap_b"] = news["anlagen_kap_b"]
        if "luftfracht_preis" in news: overrides["luftfracht_preis"] = news["luftfracht_preis"]
        if "einkauf_staffel" in news and isinstance(news["einkauf_staffel"], str):
            try:
                staffel = [[int(k.strip()), int(v.strip())] for pair in news["einkauf_staffel"].split(",") for k, v in [pair.split(":")]]
                if staffel: overrides["einkauf_staffel"] = staffel
            except Exception:
                pass
        return overrides

    def _berechne_mit_news(self, decisions, state_override, target_period):
        """Fuehrt _berechne mit temporaeren News-Parametern aus."""
        news = self.get_news(target_period)
        overrides = self._news_params_for_period(target_period)
        p = self.calibration.params
        saved = {k: p[k] for k in overrides if k in p}
        p.update(overrides)
        calc_state = state_override
        if news and "bip_wachstum" in news:
            calc_state = deepcopy(state_override or self.state)
            calc_state["pot_absatz_vor"] = calc_state.get("pot_absatz_vor", 43000) * (1 + news["bip_wachstum"] / 100.0)
        try:
            return self._berechne(decisions, state_override=calc_state)
        finally:
            p.update(saved)
            for k in overrides:
                if k not in saved:
                    del p[k]

    VERGLEICHS_KENNZAHLEN = [
        ("Aktienkurs",       "aktienkurs",       "aktienkurs"),
        ("Pot.Absatz",       "pot_absatz",        "pot_absatz"),
        ("Umsatz",           "umsatz",            "umsatz_gesamt"),
        ("EBIT",             "ebit",              "ebit"),
        ("NOPAT",            "nopat",             "nopat"),
        ("Periodenueb.",     "periodenueberschuss", "periodenueberschuss"),
        ("Eigenkapital",     "eigenkapital",      "eigenkapital"),
        ("MVA",              "mva",               "mva"),
        ("EVA",              "eva",               "eva"),
        ("KZ-Index",         "kz_index",          "kz_index"),
        ("Tech-Index",       "tech_index",        "tech_index"),
    ]

    def _build_full_decisions(self, report_data, prev_data):
        dec = report_data.get("decisions", {})
        temp_state = self._state_from_report(prev_data)
        fe_prev = prev_data.get("personal_fe_end", prev_data.get("personal_fe", 34))
        fe_curr = report_data.get("personal_fe_end", report_data.get("personal_fe", fe_prev))
        preis_m2 = dec.get("preis_m2", 0)
        werbung_m2 = dec.get("werbung_m2", 0)
        markt2_aktiv = (preis_m2 > 0 or werbung_m2 > 0) and self._is_markt2_open_for_period(report_data.get("periode", 0))
        full = {
            "preis_m1": dec.get("preis_m1", report_data.get("preis", prev_data.get("preis", 3000))),
            "werbung_m1": dec.get("werbung_m1", report_data.get("werbung", prev_data.get("werbung", 6.0))),
            "vertrieb_ma": dec.get("vertrieb_ma", report_data.get("vertrieb_ma", 100)),
            "fertigungsmenge": dec.get("fertigungsmenge", self.BASIS_KUMUL_FERTIGUNG),
            "fe_personal_aenderung": int(round(fe_curr - fe_prev)),
            "oeko_budget": 0, "ci_budget": 0, "rationalisierung": 0, "wertanalyse": 0,
            "neue_anlagen_a": int(dec.get("neue_anlagen_a", 0)),
            "neue_anlagen_b": 0,
            "personal_aenderung_fert": int(dec.get("personal_aenderung_fert", 0)),
            "grossabnehmer": int(dec.get("grossabnehmer", 0)),
            "markt2_aktiv": markt2_aktiv,
            "preis_m2": float(preis_m2) if preis_m2 else 0,
            "werbung_m2": float(werbung_m2) if werbung_m2 else 0,
            "kredit": 0, "dividende": 0,
        }
        full["neue_anlagen_c"] = dec.get("neue_anlagen_c", temp_state.get("neue_anlagen_c", 0))
        full["personal_fe"] = dec.get("personal_fe", temp_state.get("personal_fe", 35.0))
        full["personal_fe_gen2"] = dec.get("personal_fe_gen2", temp_state.get("personal_fe_gen2", 0))
        full["oeko_budget_gen1"] = dec.get("oeko_budget_gen1", temp_state.get("oeko_budget_gen1", 0.0))
        full["oeko_budget_gen2"] = dec.get("oeko_budget_gen2", temp_state.get("oeko_budget_gen2", 0.0))
        full["ci_budget"] = dec.get("ci_budget", temp_state.get("ci_budget", 0.0))
        full["grossabnehmer"] = dec.get("grossabnehmer", temp_state.get("grossabnehmer", 0))
        full["desinvest"] = dec.get("desinvest", temp_state.get("desinvest", []))
        full["einkauf_menge"] = dec.get("einkauf_menge", temp_state.get("einkauf_menge", full.get("fertigungsmenge", 0)))
        full["marktforschung_aktiv"] = dec.get("marktforschung_aktiv", False)
        full["vertrieb_m2"] = dec.get("vertrieb_m2", temp_state.get("vertrieb_m2", 0))
        return full

    def backtest(self):
        periods = self.history.all_periods()
        if len(periods) < 2:
            print("  Mindestens 2 Perioden noetig.")
            return
        tested = 0
        all_abw = {name: [] for name, _, _ in self.VERGLEICHS_KENNZAHLEN}
        print(f"\n{'='*100}")
        print(f"  BACKTEST – {len(periods)-1} Uebergaenge | {len(periods)} Perioden | {self.N_TEAMS} Teams")
        print(f"{'='*100}")
        for i in range(len(periods) - 1):
            prev_data = self.history.get_period(periods[i])
            curr_data = self.history.get_period(periods[i + 1])
            if not isinstance(prev_data, dict) or not isinstance(curr_data, dict):
                continue
            target_p = periods[i + 1]
            full_dec = self._build_full_decisions(curr_data, prev_data)
            full_dec["alle_unternehmen"] = curr_data.get("alle_unternehmen", {})
            temp_state = self._state_from_report(prev_data)
            temp_state["branche_nicht_gedeckt"] = prev_data.get("branche_nicht_gedeckt_summe", 0)
            temp_state["branche_avg_preis"] = prev_data.get("branche_avg_preis", temp_state.get("branche_avg_preis", 3000))
            erg = self._berechne_mit_news(full_dec, temp_state, target_p)
            if not isinstance(erg, dict):
                continue
            print(f"\n  P{periods[i]}->P{periods[i+1]}:")
            print(f"  {'Kennzahl':22s} {'Simuliert':>12s} {'Real':>12s} {'Abw.':>8s}  Status")
            print(f"  {'-'*22} {'-'*12} {'-'*12} {'-'*8}  {'-'*8}")
            for name, sim_key, real_key in self.VERGLEICHS_KENNZAHLEN:
                sim_val = erg.get(sim_key)
                real_val = curr_data.get(real_key)
                if sim_val is not None and real_val is not None and real_val != 0:
                    abw = (sim_val - real_val) / abs(real_val) * 100
                    all_abw[name].append(abs(abw))
                    abs_err = abs(abw)
                    if abs_err < 5:
                        status = "EXZELLENT"
                    elif abs_err < 15:
                        status = "GUT"
                    elif abs_err < 30:
                        status = "AKZEPTABEL"
                    else:
                        status = "SCHWACH"
                    print(f"  {name:22s} {sim_val:>12.2f} {real_val:>12.2f} {abw:>+7.1f}%  {status}")
            tested += 1
        if tested == 0:
            return
        print(f"\n{'='*100}")
        self._print_genauigkeit(all_abw, tested)

    def _print_genauigkeit(self, all_abw, n_perioden):
        print(f"  MODELLGENAUIGKEIT ({n_perioden} Perioden)")
        print(f"{'='*100}")
        print(f"  {'Kennzahl':22s} {'Avg':>8s} {'Max':>9s} {'Konfidenz':>10s}  Bewertung")
        print(f"  {'-'*22} {'-'*8} {'-'*9} {'-'*10}  {'-'*12}")
        total = []
        for name, _, _ in self.VERGLEICHS_KENNZAHLEN:
            vals = all_abw.get(name, [])
            if not vals:
                continue
            avg, mx = sum(vals)/len(vals), max(vals)
            k = max(0, 100 - avg * 5)
            total.append(k)
            rating = "EXZELLENT" if avg < 2 else "GUT" if avg < 5 else "BRAUCHBAR" if avg < 15 else "SCHWACH"
            print(f"  {name:22s} {avg:>7.1f}% {mx:>8.1f}% {k:>9.0f}%  {rating}")
        if total:
            gesamt = sum(total)/len(total)
            print(f"\n  {'GESAMT-KONFIDENZ':22s} {'':>8s} {'':>9s} {gesamt:>9.0f}%")
        print(f"{'='*100}")
        return total

    def modellstatus(self):
        periods = self.history.all_periods()
        n = len(periods)
        print(f"\n{'='*100}")
        print(f"  MODELLSTATUS – Eagle Eye V6.1 (Auto-Fit + Optimizer)")
        print(f"{'='*100}")
        print(f"\n  Datenbasis: {n} Perioden ({', '.join(f'P{p}' for p in periods)})")
        print(f"  Neue Mechanismen: Nacharbeit, Ueberstunden, Vertrieb->Absatz, Bekanntheit,")
        print(f"    KZ(4 Faktoren), Lagerkosten, Umweltindex+Strafe, Rating->Zins,")
        print(f"    Produktivitaet, Lernkurve, Rationalisierung, Wertanalyse, Liquiditaet 80/20")
        p = self.calibration.params
        print(f"\n  Kalibrierte Parameter:")
        for name, val, desc in [
            ("Price Elasticity", p["price_elasticity"], "Preis-Absatz"),
            ("Werbung Exp.", p["werbung_exponent"], "Werbung-Absatz"),
            ("Tech Exp.", p["tech_exponent"], "Tech-Absatz"),
            ("Vertrieb Exp.", p["vertrieb_exponent"], "Vertrieb-Absatz"),
            ("Bekanntheit Exp.", p["bekanntheit_exponent"], "Bekanntheit-Absatz"),
            ("Tech/MEUR", p["tech_per_meur"], "Tech-Index/MEUR F&E"),
            ("MVA-Faktor", p["mva_delta_faktor"], "MVA-Delta/NOPAT"),
            ("Nacharbeit %", p["nacharbeit_basis_pct"], "Basis-Nacharbeit"),
            ("Basiszins", p["basiszins"], "Zins vor Rating"),
        ]:
            print(f"    {name:22s}: {val:>8.4f}  ({desc})")
        if n >= 2:
            print(f"\n  Backtest:")
            all_abw = {name: [] for name, _, _ in self.VERGLEICHS_KENNZAHLEN}
            for i in range(len(periods) - 1):
                prev_data = self.history.get_period(periods[i])
                curr_data = self.history.get_period(periods[i + 1])
                if not isinstance(prev_data, dict) or not isinstance(curr_data, dict):
                    continue
                target_p = periods[i + 1]
                full_dec = self._build_full_decisions(curr_data, prev_data)
                full_dec["alle_unternehmen"] = curr_data.get("alle_unternehmen", {})
                temp_state = self._state_from_report(prev_data)
                temp_state["branche_nicht_gedeckt"] = prev_data.get("branche_nicht_gedeckt_summe", 0)
                temp_state["branche_avg_preis"] = prev_data.get("branche_avg_preis", 3000)
                erg = self._berechne_mit_news(full_dec, temp_state, target_p)
                if not isinstance(erg, dict):
                    continue
                for name, sim_key, real_key in self.VERGLEICHS_KENNZAHLEN:
                    sim_val = erg.get(sim_key)
                    real_val = curr_data.get(real_key)
                    if sim_val is not None and real_val is not None and real_val != 0:
                        all_abw[name].append(abs((sim_val - real_val) / real_val * 100))
            self._print_genauigkeit(all_abw, n - 1)
        print()

    def geschaetzte_abweichung(self):
        periods = self.history.all_periods()
        if len(periods) < 2:
            return {}
        all_abw = {}
        for i in range(len(periods) - 1):
            prev_data = self.history.get_period(periods[i])
            curr_data = self.history.get_period(periods[i + 1])
            if not isinstance(prev_data, dict) or not isinstance(curr_data, dict):
                continue
            target_p = periods[i + 1]
            full_dec = self._build_full_decisions(curr_data, prev_data)
            full_dec["alle_unternehmen"] = curr_data.get("alle_unternehmen", {})
            temp_state = self._state_from_report(prev_data)
            temp_state["branche_nicht_gedeckt"] = prev_data.get("branche_nicht_gedeckt_summe", 0)
            temp_state["branche_avg_preis"] = prev_data.get("branche_avg_preis", 3000)
            erg = self._berechne_mit_news(full_dec, temp_state, target_p)
            if not isinstance(erg, dict):
                continue
            for name, sim_key, real_key in self.VERGLEICHS_KENNZAHLEN:
                sim_val = erg.get(sim_key)
                real_val = curr_data.get(real_key)
                if sim_val is not None and real_val is not None and real_val != 0:
                    all_abw.setdefault(name, []).append(abs((sim_val - real_val) / real_val * 100))
        return {name: sum(v)/len(v) for name, v in all_abw.items() if v}

    def wettbewerber_analyse(self):
        periods = self.history.all_periods()
        if not periods:
            print("  Keine Daten verfuegbar.")
            return
        last_p = periods[-1]
        data = self.history.get_period(last_p)
        if not isinstance(data, dict):
            print(f"  Keine gueltigen Wettbewerberdaten in P{last_p}.")
            return
        alle = data.get("alle_unternehmen", {})
        if not alle:
            print(f"  Keine Wettbewerberdaten in P{last_p}.")
            return

        kennzahlen = [
            ("Preis", "preis", "EUR", "{:>8,.0f}"),
            ("Technologie", "tech_index", "Idx", "{:>8.1f}"),
            ("Werbung", "werbung", "MEUR", "{:>8.1f}"),
            ("Vertrieb-MA", "vertrieb_ma", "Anz", "{:>8.0f}"),
            ("KZ-Index", "kz_index", "Idx", "{:>8.1f}"),
            ("Bekanntheit", "bekanntheit", "Idx", "{:>8.1f}"),
            ("Pot.Absatz", "pot_absatz", "Stk", "{:>8,.0f}"),
            ("Tats.Absatz", "tats_absatz_markt", "Stk", "{:>8,.0f}"),
            ("Nicht gedeckt", "nicht_gedeckt", "Stk", "{:>8,.0f}"),
            ("Umsatz Markt", "umsatz_markt", "MEUR", "{:>8.1f}"),
            ("Marktanteil", "marktanteil_markt", "%", "{:>8.1f}"),
        ]

        teams = sorted(alle.keys())
        own = f"U{self.u_nr}"

        print(f"\n{'='*100}")
        print(f"  WETTBEWERBER-ANALYSE – Periode {last_p}")
        print(f"{'='*100}")
        header = f"  {'Kennzahl':15s} {'Einh.':>5s}"
        for t in teams:
            mark = " *" if t == own else ""
            header += f" {t+mark:>10s}"
        print(header)
        print(f"  {'-'*15} {'-'*5}" + "".join(f" {'-'*10}" for _ in teams))

        for name, key, unit, fmt in kennzahlen:
            line = f"  {name:15s} {unit:>5s}"
            vals = []
            for t in teams:
                v = alle[t].get(key, 0)
                vals.append(v)
                line += f" {fmt.format(v):>10s}"
            print(line)

        branche_avg = data.get("branche_avg_preis", 0)
        branche_ungedeckt = data.get("branche_nicht_gedeckt_summe", 0)
        branche_pot = data.get("branche_pot_absatz_summe", 0)
        branche_tats = data.get("branche_tats_absatz_summe", 0)

        print(f"\n  Branchen-Durchschnittspreis : {branche_avg:,.0f} EUR")
        print(f"  Branche Pot.Absatz Summe   : {branche_pot:,.0f} Stk")
        print(f"  Branche Tats.Absatz Summe  : {branche_tats:,.0f} Stk")
        print(f"  Branche Nicht Gedeckt      : {branche_ungedeckt:,.0f} Stk")

        own_data = alle.get(own, {})
        if own_data:
            own_preis = own_data.get("preis", 0)
            pos = f"{'UNTER' if own_preis < branche_avg else 'UEBER'} Durchschnitt"
            delta_pct = (own_preis - branche_avg) / max(branche_avg, 1) * 100
            print(f"\n  Unsere Position: {own_preis:,.0f} EUR ({delta_pct:+.1f}% {pos})")
            own_ma = own_data.get("marktanteil_markt", 0)
            print(f"  Unser Marktanteil: {own_ma:.1f}%")
        print(f"{'='*100}")


# ---------------------------------------------------------------------------
# CLI Interface
# ---------------------------------------------------------------------------

def main():
    sim = TOPSIM_EagleEye_V5()

    def _print_menu():
        p = sim.state["periode"]
        n = sim.history.period_count()
        nw = len(sim.news)
        nr = sim._scenario_nr
        markt2 = "offen" if sim.state.get("markt2_offen") else "geschl."
        print(f"""
╔═══════════════════════════════════════════════════════════════╗
║   TOPSIM Eagle Eye V6.1 – Auto-Fit + Optimizer                ║
║   State: P{p:<2d}  |  History: {n} Reports  |  News: {nw} Per.  |  Szen: #{nr:<3d}  ║
╠═══════════════════════════════════════════════════════════════╣
║  [1] Szenario testen (State bleibt P{p})                      ║
║  [2] Report importieren (re-kalibriert automatisch)           ║
║  [3] Backtest (alle Kennzahlen gegen echte Daten)             ║
║  [4] Kalibrierung erzwingen                                   ║
║  [5] Parameter anzeigen                                       ║
║  [6] State anzeigen                                           ║
║  [7] State auf andere Periode setzen                          ║
║  [8] Multi-Perioden-Planung                                   ║
║  [9] Modellstatus & Genauigkeit                               ║
║  [10] Wettbewerber-Analyse                                    ║
║  [11] Wirtschaftsnachrichten verwalten/anwenden               ║
║  [12] Optimale Entscheidungen berechnen                       ║
║  [0] Beenden                                                  ║
╚═══════════════════════════════════════════════════════════════╝""")

    _print_menu()

    while True:
        imported = sim._ingest_all_reports(quiet=True)
        if imported > 0:
            sim._load_state()
            print(f"\n  Auto-Import: {imported} neuer Report erkannt. State auf P{sim.state['periode']} aktualisiert.")
        choice = input("\nAuswahl [0-12]: ").strip()

        if choice == "1":
            sim.input_decisions()
            sim.simuliere_periode()
        elif choice == "2":
            path = input("  Pfad (oder Enter = reports/ scannen): ").strip()
            if path and os.path.exists(path):
                sim.ingest_report(path)
            elif not path:
                sim._ingest_all_reports()
                sim._load_state()
            else:
                print(f"  Datei nicht gefunden: {path}")
            print(f"  State jetzt auf Periode {sim.state['periode']}.")
        elif choice == "3":
            sim.backtest()
        elif choice == "4":
            sim.calibration.calibrate(sim.history)
        elif choice == "5":
            print("\n  Aktuelle Parameter:")
            for k, v in sim.calibration.params.items():
                print(f"    {k:30s}: {v}")
        elif choice == "6":
            print(f"\n  Aktueller State (Basis = P{sim.state['periode']}):")
            for k, v in sim.state.items():
                print(f"    {k:30s}: {v}")
        elif choice == "7":
            avail = sim.history.all_periods()
            print(f"  Verfuegbar: {', '.join(f'P{p}' for p in avail)}")
            pnr = _prompt_int("  Auf welche Periode? ", min_val=0)
            rd = sim.history.get_period(pnr)
            if rd:
                sim.state = sim._state_from_report(rd)
                sim._save_state()
                sim._scenario_nr = 0
                print(f"  State auf P{pnr} gesetzt.")
            else:
                print(f"  Keine Daten fuer P{pnr}.")
        elif choice == "8":
            start_p = sim.state["periode"]
            n = _prompt_int(f"  Wie viele Perioden ab P{start_p}? [3]: ", default=3, min_val=1, max_val=8)
            print(f"\n  Entscheidungen fuer {n} Perioden (P{start_p+1}–P{start_p+n}):\n")
            all_decisions, prev_dec = [], None
            for i in range(n):
                pnr = start_p + i + 1
                print(f"  --- Periode {pnr} ---")
                dp = prev_dec or {}
                d = {}
                d["preis_m1"] = _prompt_float(
                    f"    Preis [{dp.get('preis_m1', sim.state['preis_vor'])}]: ",
                    default=dp.get("preis_m1", sim.state["preis_vor"]),
                    min_val=2000,
                    max_val=7000,
                )
                d["werbung_m1"] = _prompt_float(
                    f"    Werbung [{dp.get('werbung_m1', sim.state['werbung_vor'])}]: ",
                    default=dp.get("werbung_m1", sim.state["werbung_vor"]),
                    min_val=0,
                )
                d["vertrieb_ma"] = _prompt_int(
                    f"    Vertrieb-MA [{dp.get('vertrieb_ma', sim.state['personal_vertrieb'])}]: ",
                    default=dp.get("vertrieb_ma", sim.state["personal_vertrieb"]),
                    min_val=0,
                )
                d["fertigungsmenge"] = _prompt_int(
                    f"    Fertigung [{dp.get('fertigungsmenge', 50000)}]: ",
                    default=dp.get("fertigungsmenge", 50000),
                    min_val=0,
                )
                d["fe_invest_gen1"] = _prompt_float(
                    f"    F&E [{dp.get('fe_invest_gen1', 2.5)}]: ",
                    default=dp.get("fe_invest_gen1", 2.5),
                    min_val=0,
                )
                d["fe_personal_aenderung"] = _prompt_int(
                    f"    F&E Personal +/- [{dp.get('fe_personal_aenderung', 0)}]: ",
                    default=dp.get("fe_personal_aenderung", 0),
                )
                d["oeko_budget"] = _prompt_float(
                    f"    Oeko [{dp.get('oeko_budget', 0)}]: ",
                    default=dp.get("oeko_budget", 0),
                    min_val=0,
                )
                d["ci_budget"] = _prompt_float(
                    f"    CI [{dp.get('ci_budget', 0)}]: ",
                    default=dp.get("ci_budget", 0),
                    min_val=0,
                )
                d["rationalisierung"] = _prompt_float(
                    f"    Ration. [{dp.get('rationalisierung', 0)}]: ",
                    default=dp.get("rationalisierung", 0),
                    min_val=0,
                )
                d["wertanalyse"] = _prompt_float(
                    f"    Wertanalyse [{dp.get('wertanalyse', 0)}]: ",
                    default=dp.get("wertanalyse", 0),
                    min_val=0,
                )
                d["neue_anlagen_a"] = _prompt_int(
                    f"    Anl.A [{dp.get('neue_anlagen_a', 0)}]: ",
                    default=dp.get("neue_anlagen_a", 0),
                    min_val=0,
                )
                d["neue_anlagen_b"] = _prompt_int(
                    f"    Anl.B [{dp.get('neue_anlagen_b', 0)}]: ",
                    default=dp.get("neue_anlagen_b", 0),
                    min_val=0,
                )
                d["personal_aenderung_fert"] = _prompt_int(
                    f"    Pers.+/- [{dp.get('personal_aenderung_fert', 0)}]: ",
                    default=dp.get("personal_aenderung_fert", 0),
                )
                d["grossabnehmer"] = _prompt_int(
                    f"    GA [{dp.get('grossabnehmer', 0)}]: ",
                    default=dp.get("grossabnehmer", 0),
                    min_val=0,
                )
                if sim._is_markt2_open_for_period(pnr):
                    d["markt2_aktiv"] = _prompt_yes_no(
                        "    Markt 2 aktiv? (j/n)",
                        default=dp.get("markt2_aktiv", False),
                    )
                else:
                    d["markt2_aktiv"] = False
                if d["markt2_aktiv"]:
                    d["preis_m2"] = _prompt_float(
                        f"    Preis M2 (FCU) [{dp.get('preis_m2', 28000)}]: ",
                        default=dp.get("preis_m2", 28000),
                        min_val=10000,
                        max_val=80000,
                    )
                    d["werbung_m2"] = _prompt_float(
                        f"    Werbung M2 [{dp.get('werbung_m2', 0)}]: ",
                        default=dp.get("werbung_m2", 0),
                        min_val=0,
                    )
                else:
                    d["preis_m2"] = d["werbung_m2"] = 0
                d["kredit"] = _prompt_float(
                    f"    Kredit [{dp.get('kredit', 0)}]: ",
                    default=dp.get("kredit", 0),
                    min_val=0,
                )
                d["dividende"] = _prompt_float(
                    f"    Dividende [{dp.get('dividende', 0)}]: ",
                    default=dp.get("dividende", 0),
                    min_val=0,
                )
                all_decisions.append(d)
                prev_dec = d
                print()
            ergebnisse = sim.simuliere_multi(all_decisions)
            print(f"\n{'='*120}")
            print(f"  MULTI-PERIODEN P{start_p+1}–P{start_p+n}")
            print(f"{'='*120}")
            print(f"  {'P':>4s} | {'Aktie':>7s} | {'Umsatz':>8s} | {'EBIT':>8s} | {'EK':>8s} | {'MVA':>8s} | {'Pot.Abs':>8s} | {'Lager':>7s} | {'KZ':>5s} | {'Umwelt':>6s}")
            for erg in ergebnisse:
                print(f"  P{erg['periode']:>2d} | {erg['aktienkurs']:>7.0f} | {erg['umsatz']:>8.2f} | {erg['ebit']:>8.2f} | {erg['eigenkapital']:>8.2f} | {erg['mva']:>8.2f} | {erg['pot_absatz']:>8.0f} | {erg['neues_lager']:>7.0f} | {erg['kz_index']:>5.1f} | {erg['umweltindex']:>6.1f}")
            print(f"{'='*120}")
            for erg in ergebnisse:
                sim._print_abweichung(erg)
            exp = _prompt_yes_no("  Exportieren? (j/n)", default=False)
            if exp:
                fname = input(f"  Dateiname [szenario_P{start_p+1}-P{start_p+n}.csv]: ").strip() or f"szenario_P{start_p+1}-P{start_p+n}.csv"
                sim.export_ergebnisse(ergebnisse, fname)
        elif choice == "9":
            sim.modellstatus()
        elif choice == "10":
            sim.wettbewerber_analyse()
        elif choice == "11":
            print(f"\n  Wirtschaftsnachrichten (gespeichert: {', '.join(f'P{k}' for k in sorted(sim.news.keys()))})")
            sub = input("  [a] Anzeigen | [e] Eingeben/Bearbeiten | [w] Anwenden: ").strip().lower()
            if sub == "a":
                pn = _prompt_int(
                    f"  Welche Periode? [{sim.state['periode']+1}]: ",
                    default=sim.state["periode"] + 1,
                    min_val=0,
                )
                sim.print_news(pn)
            elif sub == "e":
                pn = _prompt_int(
                    f"  Fuer welche Periode? [{sim.state['periode']+1}]: ",
                    default=sim.state["periode"] + 1,
                    min_val=0,
                )
                existing = sim.get_news(pn)
                nd = dict(existing) if existing else {}
                NEWS_FELDER = [
                    ("lohn_einkauf", "Lohn Einkauf (TEUR)"),
                    ("lohn_verwaltung", "Lohn Verwaltung (TEUR)"),
                    ("lohn_fertigung", "Lohn Fertigung (TEUR)"),
                    ("lohn_fe", "Lohn F&E (TEUR)"),
                    ("lohn_vertrieb", "Lohn Vertrieb (TEUR)"),
                    ("betriebsstoff", "Betriebsstoffe (EUR/Stk)"),
                    ("basiszins", "Zinsen Ueberziehungskr. (%)"),
                    ("transport_m1", "Transport Markt 1 (EUR/Stk)"),
                    ("transport_m2", "Transport Markt 2 (EUR/Stk)"),
                    ("einstellungskosten", "Einstellungskosten (TEUR/MA)"),
                    ("entlassungskosten", "Entlassungskosten (TEUR/MA)"),
                    ("anlagen_kap_b", "Kapazität Typ B (Stk)"),
                    ("luftfracht_preis", "Luftfracht Preis (EUR/Stk)"),
                    ("einkauf_staffel", "Einkaufsstaffel (Grenze:Preis, ...)"),
                    ("wechselkurs", "Wechselkurs (EUR/FCU)"),
                    ("anlagen_preis_b", "Kaufpreis Typ B (MEUR)"),
                    ("anlagen_fix_b", "Fixkosten Typ B (MEUR)"),
                    ("bip_wachstum", "BIP-Wachstum (%)"),
                    ("markt2_offen", "Markt 2 offen? (ja/nein)"),
                ]
                print(f"\n  News P{pn} bearbeiten (Enter = Wert behalten):")
                for key, label in NEWS_FELDER:
                    curr = nd.get(key, "")
                    val = input(f"    {label} [{curr}]: ").strip()
                    if val:
                        if key == "markt2_offen":
                            nd[key] = val.lower() in {"1", "true", "ja", "yes", "y", "j"}
                        elif key == "einkauf_staffel":
                            nd[key] = val
                        else:
                            try:
                                nd[key] = float(val)
                            except ValueError:
                                print(f"    Ungueltiger numerischer Wert fuer {label}, Feld bleibt unveraendert.")
                sim.set_news(pn, nd)
                sim.print_news(pn)
            elif sub == "w":
                pn = _prompt_int(
                    f"  Welche Periode anwenden? [{sim.state['periode']+1}]: ",
                    default=sim.state["periode"] + 1,
                    min_val=0,
                )
                sim.apply_news(pn)
        elif choice == "12":
            print(f"\n  Optimierung fuer P{sim.state['periode']+1}")
            print("  Ziel: [1] Balanced (Mix) | [2] Aktienkurs | [3] EBIT | [4] Eigenkapital | [5] Vergleich alle")
            z = input("  Wahl [1]: ").strip() or "1"
            pm = input("  Gegner-Prognose [1] MCI-Advanced | [2] Simple Trend [1]: ").strip() or "1"
            pred_mode = "advanced" if pm == "1" else "simple"
            ziel_map = {"1": "balanced", "2": "aktienkurs", "3": "ebit", "4": "eigenkapital"}
            if z == "5":
                sim.optimiere_vergleich()
            else:
                ziel = ziel_map.get(z, "balanced")
                result = sim.optimiere_entscheidungen(ziel=ziel, pred_mode=pred_mode)
                if result:
                    use = _prompt_yes_no("\n  Diese Werte als Entscheidungen uebernehmen? (j/n)", default=False)
                    if use:
                        sim.decisions = result[0]
                        print("  Entscheidungen uebernommen.")
        elif choice == "0":
            print("\n  Eagle Eye V6.1 beendet.")
            break
        elif choice == "m":
            _print_menu()


if __name__ == "__main__":
    main()
